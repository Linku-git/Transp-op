from __future__ import annotations

import uuid
from datetime import date
from io import BytesIO

import openpyxl
import pytest
from httpx import AsyncClient

from tests.conftest import login_as_admin


# ---------------------------------------------------------------------------
# Workbook builder helpers
# ---------------------------------------------------------------------------


def _create_test_workbook(
    include_sites: bool = True,
    include_effectif: bool = True,
    include_usages: bool = True,
    include_contraintes: bool = True,
    include_parc: bool = True,
    include_absences: bool = True,
    site_code: str | None = None,
    matricule: str | None = None,
    site_lat: float = 33.57,
    site_lng: float = -7.59,
    emp_lat: float = 33.58,
    emp_lng: float = -7.60,
    extra_site_rows: list[list] | None = None,
    extra_effectif_rows: list[list] | None = None,
    extra_absence_rows: list[list] | None = None,
) -> bytes:
    """Build a minimal valid multi-sheet Excel template in memory.

    Returns raw bytes suitable for uploading via the API.
    """
    if site_code is None:
        site_code = f"XS-{uuid.uuid4().hex[:6]}"
    if matricule is None:
        matricule = f"XE-{uuid.uuid4().hex[:6]}"

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    if include_sites:
        ws = wb.create_sheet("SITES")
        ws.append(["Code Site", "Nom Site", "Adresse", "Ville", "Latitude", "Longitude", "Nombre Equipes"])
        ws.append([site_code, "Site Alpha", "10 Rue Test", "Casablanca", site_lat, site_lng, 2])
        if extra_site_rows:
            for row in extra_site_rows:
                ws.append(row)

    if include_effectif:
        ws = wb.create_sheet("EFFECTIF")
        ws.append(["Matricule", "Prenom", "Nom", "Code Site", "Equipe", "Adresse", "Ville", "Latitude", "Longitude", "PMR", "Departement", "Telephone", "Mode Transport", "Opt-in"])
        ws.append([matricule, "Ahmed", "Benali", site_code, "Equipe 1", "20 Rue Employé", "Casablanca", emp_lat, emp_lng, "Non", "IT", "0600000001", "Voiture", "Oui"])
        if extra_effectif_rows:
            for row in extra_effectif_rows:
                ws.append(row)

    if include_usages:
        ws = wb.create_sheet("USAGES & MODES")
        ws.append(["Matricule", "Mode Principal", "Mode Alternatif", "Distance km", "Temps Trajet", "Frequence", "Accepte Point Commun"])
        ws.append([matricule, "Voiture", "Bus", 15.5, 30, "Quotidien", "Oui"])

    if include_contraintes:
        ws = wb.create_sheet("CONTRAINTES")
        ws.append(["Parametre", "Valeur"])
        ws.append(["max_detour_minutes", "15"])
        ws.append(["max_vehicle_capacity", "50"])

    if include_parc:
        ws = wb.create_sheet("PARC EXISTANT")
        ws.append(["Type Vehicule", "Capacite", "Immatriculation"])
        ws.append(["Minibus", 20, "AB-123-CD"])
        ws.append(["Bus", 50, "EF-456-GH"])

    if include_absences:
        ws = wb.create_sheet("ABSENCES")
        ws.append(["Matricule", "Type Absence", "Date Debut", "Date Fin"])
        ws.append([matricule, "vacation", date(2026, 7, 1), date(2026, 7, 15)])
        if extra_absence_rows:
            for row in extra_absence_rows:
                ws.append(row)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# API helpers
# ---------------------------------------------------------------------------


async def _create_site(client: AsyncClient, token: str, code: str = "SITE-A") -> str:
    """Create a site via the API and return its UUID."""
    resp = await client.post(
        "/api/v1/sites/",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "code": code,
            "name": f"Site {code}",
            "address": "10 Rue Test",
            "city": "Casablanca",
            "lat": 33.57,
            "lng": -7.59,
        },
    )
    assert resp.status_code == 201, f"Site creation failed: {resp.text}"
    return resp.json()["id"]


async def _create_employee(
    client: AsyncClient, token: str, site_id: str, matricule: str = "EMP-001"
) -> str:
    """Create an employee via the API and return its UUID."""
    resp = await client.post(
        "/api/v1/employees/",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "matricule": matricule,
            "first_name": "Ahmed",
            "last_name": "Benali",
            "site_id": site_id,
            "city": "Casablanca",
            "lat": 33.58,
            "lng": -7.60,
        },
    )
    assert resp.status_code == 201, f"Employee creation failed: {resp.text}"
    return resp.json()["id"]


async def _upload_excel(
    client: AsyncClient,
    token: str,
    file_bytes: bytes,
    endpoint: str = "/api/v1/import/excel",
    params: dict | None = None,
) -> dict:
    """Upload an Excel file and return the JSON response body."""
    resp = await client.post(
        endpoint,
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("template.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        params=params,
    )
    return {"status_code": resp.status_code, "body": resp.json()}


def _find_sheet(body: dict, sheet_name: str) -> dict | None:
    """Find a SheetResult in the response by name."""
    for s in body.get("sheets", []):
        if s["sheet"] == sheet_name:
            return s
    return None


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_parse_valid_template(client: AsyncClient) -> None:
    """Upload a full template — all 6 sheets are processed successfully."""
    token = await login_as_admin(client)
    code = f"XV-{uuid.uuid4().hex[:6]}"
    mat = f"XM-{uuid.uuid4().hex[:6]}"

    file_bytes = _create_test_workbook(site_code=code, matricule=mat)
    result = await _upload_excel(client, token, file_bytes)

    assert result["status_code"] == 200, f"Unexpected: {result['body']}"
    body = result["body"]
    assert body["is_preview"] is False

    # All 6 sheets should appear in the results
    sheet_names = {s["sheet"] for s in body["sheets"]}
    assert "SITES" in sheet_names
    assert "EFFECTIF" in sheet_names
    assert "USAGES & MODES" in sheet_names
    assert "CONTRAINTES" in sheet_names
    assert "PARC EXISTANT" in sheet_names
    assert "ABSENCES" in sheet_names


@pytest.mark.asyncio
async def test_parse_sites_sheet(client: AsyncClient) -> None:
    """SITES sheet creates site records in the database."""
    token = await login_as_admin(client)
    unique_code = f"XS-{uuid.uuid4().hex[:6]}"

    file_bytes = _create_test_workbook(
        include_effectif=False,
        include_usages=False,
        include_contraintes=False,
        include_parc=False,
        include_absences=False,
        site_code=unique_code,
    )
    result = await _upload_excel(client, token, file_bytes)

    assert result["status_code"] == 200
    sites_sheet = _find_sheet(result["body"], "SITES")
    assert sites_sheet is not None
    assert sites_sheet["rows_imported"] == 1
    assert sites_sheet["rows_skipped"] == 0
    assert len(sites_sheet["errors"]) == 0

    # Verify site was actually created by fetching by code
    code_resp = await client.get(
        f"/api/v1/sites/code/{unique_code}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert code_resp.status_code == 200
    assert code_resp.json()["code"] == unique_code


@pytest.mark.asyncio
async def test_parse_effectif_sheet(client: AsyncClient) -> None:
    """EFFECTIF sheet creates employee records linked to existing site."""
    token = await login_as_admin(client)
    site_code = f"ES-{uuid.uuid4().hex[:6]}"
    matricule = f"EM-{uuid.uuid4().hex[:6]}"

    # Pre-create the site so the FK lookup succeeds
    await _create_site(client, token, code=site_code)

    file_bytes = _create_test_workbook(
        include_sites=False,
        include_usages=False,
        include_contraintes=False,
        include_parc=False,
        include_absences=False,
        site_code=site_code,
        matricule=matricule,
    )
    result = await _upload_excel(client, token, file_bytes)

    assert result["status_code"] == 200
    effectif_sheet = _find_sheet(result["body"], "EFFECTIF")
    assert effectif_sheet is not None
    assert effectif_sheet["rows_imported"] == 1
    assert len(effectif_sheet["errors"]) == 0

    # Verify employee was created
    emp_resp = await client.get(
        "/api/v1/employees/",
        headers={"Authorization": f"Bearer {token}"},
        params={"q": matricule},
    )
    assert emp_resp.status_code == 200
    employees = emp_resp.json()["data"]
    assert len(employees) >= 1
    found = [e for e in employees if e["matricule"] == matricule]
    assert len(found) == 1
    assert found[0]["first_name"] == "Ahmed"
    assert found[0]["last_name"] == "Benali"


@pytest.mark.asyncio
async def test_parse_usages_sheet(client: AsyncClient) -> None:
    """USAGES & MODES sheet is parsed and validated without error (no DB write yet)."""
    token = await login_as_admin(client)
    site_code = f"US-{uuid.uuid4().hex[:6]}"
    matricule = f"UM-{uuid.uuid4().hex[:6]}"

    # Create site and employee so matricule lookup succeeds
    site_id = await _create_site(client, token, code=site_code)
    await _create_employee(client, token, site_id, matricule=matricule)

    file_bytes = _create_test_workbook(
        include_sites=False,
        include_effectif=False,
        include_contraintes=False,
        include_parc=False,
        include_absences=False,
        site_code=site_code,
        matricule=matricule,
    )
    result = await _upload_excel(client, token, file_bytes)

    assert result["status_code"] == 200
    usages_sheet = _find_sheet(result["body"], "USAGES & MODES")
    assert usages_sheet is not None
    assert usages_sheet["rows_imported"] == 1
    assert len(usages_sheet["errors"]) == 0


@pytest.mark.asyncio
async def test_parse_contraintes_sheet(client: AsyncClient) -> None:
    """CONTRAINTES sheet is parsed and validated without error (no DB write yet)."""
    token = await login_as_admin(client)

    file_bytes = _create_test_workbook(
        include_sites=False,
        include_effectif=False,
        include_usages=False,
        include_parc=False,
        include_absences=False,
    )
    result = await _upload_excel(client, token, file_bytes)

    assert result["status_code"] == 200
    contraintes_sheet = _find_sheet(result["body"], "CONTRAINTES")
    assert contraintes_sheet is not None
    assert contraintes_sheet["rows_imported"] == 2  # two key-value rows
    assert len(contraintes_sheet["errors"]) == 0


@pytest.mark.asyncio
async def test_parse_parc_sheet(client: AsyncClient) -> None:
    """PARC EXISTANT sheet is parsed and validated without error (no DB write yet)."""
    token = await login_as_admin(client)

    file_bytes = _create_test_workbook(
        include_sites=False,
        include_effectif=False,
        include_usages=False,
        include_contraintes=False,
        include_absences=False,
    )
    result = await _upload_excel(client, token, file_bytes)

    assert result["status_code"] == 200
    parc_sheet = _find_sheet(result["body"], "PARC EXISTANT")
    assert parc_sheet is not None
    assert parc_sheet["rows_imported"] == 2  # Minibus + Bus
    assert len(parc_sheet["errors"]) == 0


@pytest.mark.asyncio
async def test_parse_absences_sheet(client: AsyncClient) -> None:
    """ABSENCES sheet creates leave records linked to existing employees."""
    token = await login_as_admin(client)
    site_code = f"AB-{uuid.uuid4().hex[:6]}"
    matricule = f"AM-{uuid.uuid4().hex[:6]}"

    # Create site and employee prerequisites
    site_id = await _create_site(client, token, code=site_code)
    emp_id = await _create_employee(client, token, site_id, matricule=matricule)

    file_bytes = _create_test_workbook(
        include_sites=False,
        include_effectif=False,
        include_usages=False,
        include_contraintes=False,
        include_parc=False,
        site_code=site_code,
        matricule=matricule,
    )
    result = await _upload_excel(client, token, file_bytes)

    assert result["status_code"] == 200
    absences_sheet = _find_sheet(result["body"], "ABSENCES")
    assert absences_sheet is not None
    assert absences_sheet["rows_imported"] == 1
    assert len(absences_sheet["errors"]) == 0

    # Verify the leave record was created
    leave_resp = await client.get(
        f"/api/v1/leaves/?employee_id={emp_id}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert leave_resp.status_code == 200
    leaves = leave_resp.json()["data"]
    assert len(leaves) >= 1
    assert leaves[0]["leave_type"] == "vacation"


@pytest.mark.asyncio
async def test_validation_required_fields(client: AsyncClient) -> None:
    """Missing required fields are reported as errors with correct row/column."""
    token = await login_as_admin(client)

    # Build a workbook with SITES sheet missing required fields
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("SITES")
    ws.append(["Code Site", "Nom Site", "Adresse", "Ville", "Latitude", "Longitude"])
    # Row 2: missing name and city
    ws.append(["CODE-X", None, "10 Rue Test", None, 33.57, -7.59])
    # Row 3: missing code
    ws.append([None, "Site B", "20 Rue Test", "Rabat", 34.02, -6.84])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    file_bytes = buf.read()

    result = await _upload_excel(client, token, file_bytes)

    assert result["status_code"] == 200
    body = result["body"]
    assert body["total_errors"] >= 2

    sites_sheet = _find_sheet(body, "SITES")
    assert sites_sheet is not None
    assert sites_sheet["rows_skipped"] >= 2
    assert sites_sheet["rows_imported"] == 0

    # Check that specific required-field errors are present
    error_columns = [e["column"] for e in sites_sheet["errors"]]
    assert "Nom Site" in error_columns
    assert "Ville" in error_columns
    assert "Code Site" in error_columns


@pytest.mark.asyncio
async def test_validation_invalid_data_types(client: AsyncClient) -> None:
    """Invalid lat/lng values are reported as errors."""
    token = await login_as_admin(client)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("SITES")
    ws.append(["Code Site", "Nom Site", "Adresse", "Ville", "Latitude", "Longitude"])
    # Row 2: latitude out of range
    ws.append(["BAD-1", "Bad Site", "123 Street", "City", 999.0, -7.59])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    file_bytes = buf.read()

    result = await _upload_excel(client, token, file_bytes)

    assert result["status_code"] == 200
    sites_sheet = _find_sheet(result["body"], "SITES")
    assert sites_sheet is not None
    assert len(sites_sheet["errors"]) >= 1
    assert any("latitude" in e["message"].lower() or "Latitude" in e["column"] for e in sites_sheet["errors"])


@pytest.mark.asyncio
async def test_validation_invalid_site_ref(client: AsyncClient) -> None:
    """EFFECTIF rows referencing unknown site codes are reported as errors."""
    token = await login_as_admin(client)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("EFFECTIF")
    ws.append(["Matricule", "Prenom", "Nom", "Code Site", "Equipe", "Adresse", "Ville", "Latitude", "Longitude", "PMR", "Departement", "Telephone", "Mode Transport", "Opt-in"])
    ws.append(["EMP-GHOST", "Karim", "Ziani", "NONEXISTENT-SITE", "Equipe 1", "1 Rue X", "Rabat", 34.02, -6.84, "Non", "HR", "0600000002", "Bus", "Non"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    file_bytes = buf.read()

    result = await _upload_excel(client, token, file_bytes)

    assert result["status_code"] == 200
    effectif_sheet = _find_sheet(result["body"], "EFFECTIF")
    assert effectif_sheet is not None
    assert len(effectif_sheet["errors"]) >= 1
    assert effectif_sheet["rows_imported"] == 0
    assert any("NONEXISTENT-SITE" in e["message"] for e in effectif_sheet["errors"])


@pytest.mark.asyncio
async def test_preview_mode(client: AsyncClient) -> None:
    """Preview mode returns validation results without creating DB records."""
    token = await login_as_admin(client)
    site_code = f"PV-{uuid.uuid4().hex[:6]}"

    file_bytes = _create_test_workbook(
        include_effectif=False,
        include_usages=False,
        include_contraintes=False,
        include_parc=False,
        include_absences=False,
        site_code=site_code,
    )
    result = await _upload_excel(
        client, token, file_bytes, endpoint="/api/v1/import/excel/preview"
    )

    assert result["status_code"] == 200
    body = result["body"]
    assert body["is_preview"] is True

    sites_sheet = _find_sheet(body, "SITES")
    assert sites_sheet is not None
    assert sites_sheet["rows_imported"] == 1  # validated but not committed

    # Verify the site was NOT actually created
    list_resp = await client.get(
        "/api/v1/sites/",
        headers={"Authorization": f"Bearer {token}"},
    )
    site_codes = [s["code"] for s in list_resp.json()["data"]]
    assert site_code not in site_codes


@pytest.mark.asyncio
async def test_incremental_import(client: AsyncClient) -> None:
    """A second import updates existing records rather than duplicating them."""
    token = await login_as_admin(client)
    site_code = f"INC-{uuid.uuid4().hex[:6]}"

    # First import: create site
    file_bytes_1 = _create_test_workbook(
        include_effectif=False,
        include_usages=False,
        include_contraintes=False,
        include_parc=False,
        include_absences=False,
        site_code=site_code,
    )
    r1 = await _upload_excel(client, token, file_bytes_1)
    assert r1["status_code"] == 200

    # Get original site name
    code_resp = await client.get(
        f"/api/v1/sites/code/{site_code}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert code_resp.status_code == 200
    assert code_resp.json()["name"] == "Site Alpha"

    # Second import: same code but modified name
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("SITES")
    ws.append(["Code Site", "Nom Site", "Adresse", "Ville", "Latitude", "Longitude", "Nombre Equipes"])
    ws.append([site_code, "Site Alpha Updated", "10 Rue Test", "Casablanca", 33.57, -7.59, 3])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    r2 = await _upload_excel(client, token, buf.read())
    assert r2["status_code"] == 200

    sites_sheet = _find_sheet(r2["body"], "SITES")
    assert sites_sheet is not None
    assert sites_sheet["rows_imported"] == 1
    assert len(sites_sheet["errors"]) == 0

    # Verify update — the name changed
    code_resp2 = await client.get(
        f"/api/v1/sites/code/{site_code}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert code_resp2.status_code == 200
    assert code_resp2.json()["name"] == "Site Alpha Updated"


@pytest.mark.asyncio
async def test_single_sheet_import(client: AsyncClient) -> None:
    """The sheet_name query parameter imports only the specified sheet."""
    token = await login_as_admin(client)

    file_bytes = _create_test_workbook(
        include_effectif=False,
        include_usages=False,
        include_absences=False,
    )
    result = await _upload_excel(
        client,
        token,
        file_bytes,
        endpoint="/api/v1/import/excel/sheet",
        params={"sheet_name": "CONTRAINTES"},
    )

    assert result["status_code"] == 200
    body = result["body"]

    # Only the requested sheet should appear in results
    sheet_names = [s["sheet"] for s in body["sheets"]]
    assert sheet_names == ["CONTRAINTES"]

    contraintes_sheet = _find_sheet(body, "CONTRAINTES")
    assert contraintes_sheet is not None
    assert contraintes_sheet["rows_imported"] == 2
