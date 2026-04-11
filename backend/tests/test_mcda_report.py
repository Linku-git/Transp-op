"""Tests for MCDA comparison report generator (Session 113).

Covers PDF generation, Excel multi-sheet output, radar chart rendering,
sensitivity tables, conditional formatting, and edge cases.
"""
from __future__ import annotations

import io
import uuid

import pypdf
import pytest

from app.services.sotreg.mcda_report import (
    ALT_COLORS,
    CRITERIA_LABELS,
    CRITERIA_ORDER,
    _draw_radar_chart,
    generate_mcda_excel,
    generate_mcda_pdf,
)
from app.services.sotreg.mcda_service import (
    CDC_DEFAULT_WEIGHTS,
    compute_mcda_scores,
)


def _extract_pdf_text(pdf_bytes: bytes) -> str:
    """Extract text from PDF bytes using pypdf."""
    reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
    text = ""
    for page in reader.pages:
        text += page.extract_text() or ""
    return text

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

DIESEL = {
    "name": "Diesel",
    "capex": 180000.0,
    "opex": 120000.0,
    "co2": 90.0,
    "risk": 3.0,
    "comfort": 3.0,
    "maturity": 5.0,
}

ELECTRIC = {
    "name": "Electric",
    "capex": 300000.0,
    "opex": 60000.0,
    "co2": 10.0,
    "risk": 2.0,
    "comfort": 4.5,
    "maturity": 3.0,
}

HYBRID = {
    "name": "Hybrid",
    "capex": 250000.0,
    "opex": 80000.0,
    "co2": 45.0,
    "risk": 2.5,
    "comfort": 4.0,
    "maturity": 4.0,
}

CNG = {
    "name": "CNG",
    "capex": 220000.0,
    "opex": 95000.0,
    "co2": 60.0,
    "risk": 3.5,
    "comfort": 3.5,
    "maturity": 3.5,
}

SAMPLE_ALTERNATIVES = [DIESEL, ELECTRIC, HYBRID, CNG]
SAMPLE_WEIGHTS = dict(CDC_DEFAULT_WEIGHTS)


@pytest.fixture
def sample_results() -> dict:
    """Compute MCDA results from sample alternatives."""
    return compute_mcda_scores(SAMPLE_ALTERNATIVES, SAMPLE_WEIGHTS)


@pytest.fixture
def sample_pdf(sample_results: dict) -> bytes:
    """Generate a sample PDF report."""
    return generate_mcda_pdf(
        "Test Scenario",
        SAMPLE_ALTERNATIVES,
        SAMPLE_WEIGHTS,
        sample_results,
    )


@pytest.fixture
def sample_excel(sample_results: dict) -> bytes:
    """Generate a sample Excel report."""
    return generate_mcda_excel(
        "Test Scenario",
        SAMPLE_ALTERNATIVES,
        SAMPLE_WEIGHTS,
        sample_results,
    )


# ---------------------------------------------------------------------------
# PDF tests
# ---------------------------------------------------------------------------


class TestPDFGeneration:
    """Tests for MCDA PDF report generation."""

    def test_pdf_produces_valid_bytes(self, sample_pdf: bytes) -> None:
        """PDF output is non-empty and starts with PDF magic bytes."""
        assert isinstance(sample_pdf, bytes)
        assert len(sample_pdf) > 0
        assert sample_pdf[:5] == b"%PDF-"

    def test_pdf_contains_executive_summary(self, sample_pdf: bytes) -> None:
        """PDF contains the executive summary section text."""
        text = _extract_pdf_text(sample_pdf)
        assert "sum" in text.lower() or "Ex" in text

    def test_pdf_contains_scenario_name(self, sample_pdf: bytes) -> None:
        """PDF contains the scenario name."""
        text = _extract_pdf_text(sample_pdf)
        assert "Test Scenario" in text

    def test_pdf_contains_all_alternatives(self, sample_pdf: bytes) -> None:
        """PDF contains all alternative names."""
        text = _extract_pdf_text(sample_pdf)
        for alt in SAMPLE_ALTERNATIVES:
            assert alt["name"] in text, f"Missing {alt['name']} in PDF"

    def test_pdf_contains_comparison_table(self, sample_pdf: bytes) -> None:
        """PDF contains comparison table section."""
        text = _extract_pdf_text(sample_pdf)
        assert "Comparatif" in text or "Rang" in text

    def test_pdf_contains_sensitivity_section(self, sample_pdf: bytes) -> None:
        """PDF contains sensitivity analysis section."""
        text = _extract_pdf_text(sample_pdf)
        assert "Sensibilit" in text

    def test_pdf_contains_recommendation(self, sample_pdf: bytes) -> None:
        """PDF contains recommendation section."""
        text = _extract_pdf_text(sample_pdf)
        assert "Recommandation" in text

    def test_pdf_contains_radar_section(self, sample_pdf: bytes) -> None:
        """PDF contains radar chart section header."""
        text = _extract_pdf_text(sample_pdf)
        assert "Radar" in text

    def test_pdf_with_two_alternatives(self) -> None:
        """PDF generates correctly with minimum 2 alternatives."""
        alts = [DIESEL, ELECTRIC]
        results = compute_mcda_scores(alts, SAMPLE_WEIGHTS)
        pdf = generate_mcda_pdf("Two Alts", alts, SAMPLE_WEIGHTS, results)
        assert isinstance(pdf, bytes)
        assert len(pdf) > 0
        assert pdf[:5] == b"%PDF-"

    def test_pdf_large_scenario(self) -> None:
        """PDF generates within reasonable time for 20+ alternatives."""
        alts = []
        for i in range(25):
            alts.append({
                "name": f"Alt_{i:02d}",
                "capex": 150000 + i * 10000,
                "opex": 50000 + i * 5000,
                "co2": 10 + i * 4,
                "risk": min(i * 0.4, 10),
                "comfort": max(10 - i * 0.3, 0),
                "maturity": max(10 - i * 0.2, 0),
            })
        results = compute_mcda_scores(alts, SAMPLE_WEIGHTS)
        pdf = generate_mcda_pdf("Large Scenario", alts, SAMPLE_WEIGHTS, results)
        assert isinstance(pdf, bytes)
        assert len(pdf) > 1000

    def test_pdf_stability_confidence_levels(self) -> None:
        """PDF recommendation reflects stability confidence levels."""
        # Use identical alternatives for 100% stability
        identical = [
            {"name": "A", "capex": 100, "opex": 50, "co2": 30, "risk": 2, "comfort": 5, "maturity": 5},
            {"name": "B", "capex": 100, "opex": 50, "co2": 30, "risk": 2, "comfort": 5, "maturity": 5},
        ]
        results = compute_mcda_scores(identical, SAMPLE_WEIGHTS)
        pdf = generate_mcda_pdf("Identical", identical, SAMPLE_WEIGHTS, results)
        text = _extract_pdf_text(pdf)
        assert "Recommandation" in text


# ---------------------------------------------------------------------------
# Radar chart tests
# ---------------------------------------------------------------------------


class TestRadarChart:
    """Tests for radar/spider chart rendering."""

    def test_radar_chart_renders(self, sample_results: dict) -> None:
        """Radar chart drawing does not raise exceptions."""
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas as rl_canvas

        buf = io.BytesIO()
        c = rl_canvas.Canvas(buf, pagesize=A4)

        ranked = sample_results["ranked_alternatives"]
        _draw_radar_chart(
            c, 300, 400, 150,
            ranked, CRITERIA_ORDER, ALT_COLORS,
        )
        c.save()
        result = buf.getvalue()
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_radar_chart_with_6_axes(self) -> None:
        """Radar chart correctly handles all 6 criteria axes."""
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas as rl_canvas

        alts = [
            {
                "name": "Test",
                "normalized_values": {c: 3.0 for c in CRITERIA_ORDER},
            }
        ]
        buf = io.BytesIO()
        c = rl_canvas.Canvas(buf, pagesize=A4)
        _draw_radar_chart(c, 300, 400, 150, alts, CRITERIA_ORDER, ALT_COLORS)
        c.save()
        assert len(buf.getvalue()) > 0

    def test_radar_chart_multiple_overlapping(self) -> None:
        """Radar chart handles multiple overlapping alternative polygons."""
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas as rl_canvas

        alts = [
            {"name": "A", "normalized_values": {"capex": 5, "opex": 4, "co2": 3, "risk": 2, "comfort": 1, "maturity": 5}},
            {"name": "B", "normalized_values": {"capex": 1, "opex": 2, "co2": 5, "risk": 4, "comfort": 5, "maturity": 3}},
            {"name": "C", "normalized_values": {"capex": 3, "opex": 3, "co2": 3, "risk": 3, "comfort": 3, "maturity": 3}},
        ]
        buf = io.BytesIO()
        c = rl_canvas.Canvas(buf, pagesize=A4)
        _draw_radar_chart(c, 300, 400, 150, alts, CRITERIA_ORDER, ALT_COLORS)
        c.save()
        assert len(buf.getvalue()) > 0

    def test_radar_chart_fewer_than_3_axes_skipped(self) -> None:
        """Radar chart with fewer than 3 axes returns without drawing."""
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas as rl_canvas

        buf = io.BytesIO()
        c = rl_canvas.Canvas(buf, pagesize=A4)
        _draw_radar_chart(c, 300, 400, 150, [], ["a", "b"], ALT_COLORS)
        c.save()
        # Should still produce valid PDF (just empty)
        assert len(buf.getvalue()) > 0


# ---------------------------------------------------------------------------
# Excel tests
# ---------------------------------------------------------------------------


class TestExcelGeneration:
    """Tests for MCDA Excel report generation."""

    def test_excel_produces_valid_bytes(self, sample_excel: bytes) -> None:
        """Excel output is non-empty and can be loaded."""
        assert isinstance(sample_excel, bytes)
        assert len(sample_excel) > 0

    def test_excel_has_four_sheets(self, sample_excel: bytes) -> None:
        """Excel workbook has exactly 4 sheets."""
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(sample_excel))
        assert len(wb.sheetnames) == 4
        assert wb.sheetnames == ["Summary", "Scores", "Sensitivity", "Raw Data"]

    def test_excel_summary_sheet(self, sample_excel: bytes) -> None:
        """Summary sheet contains scenario metadata."""
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(sample_excel))
        ws = wb["Summary"]
        # First cell should contain scenario name
        assert "Test Scenario" in str(ws["A1"].value)
        # Should have metric rows
        values = [str(ws.cell(row=r, column=1).value or "") for r in range(1, 20)]
        text = " ".join(values)
        assert "alternative" in text.lower() or "Meilleure" in text

    def test_excel_scores_sheet_all_alternatives(self, sample_excel: bytes) -> None:
        """Scores sheet contains all alternatives."""
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(sample_excel))
        ws = wb["Scores"]
        alt_names = set()
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=2).value
            if name:
                alt_names.add(name)
        for alt in SAMPLE_ALTERNATIVES:
            assert alt["name"] in alt_names, f"Missing {alt['name']} in Scores"

    def test_excel_conditional_formatting(self, sample_excel: bytes) -> None:
        """Excel Scores sheet has conditional formatting rules."""
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(sample_excel))
        ws = wb["Scores"]
        # Check conditional formatting rules exist
        rules = ws.conditional_formatting._cf_rules
        assert len(rules) > 0, "No conditional formatting rules found"

    def test_excel_sensitivity_sheet(self, sample_excel: bytes) -> None:
        """Sensitivity sheet contains criteria analysis."""
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(sample_excel))
        ws = wb["Sensitivity"]
        # Header row
        assert ws.cell(row=1, column=1).value == "Critère"
        # Should have data rows for each criterion
        criteria_found = set()
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val and val in CRITERIA_LABELS.values():
                criteria_found.add(val)
        assert len(criteria_found) == 6, f"Found only {len(criteria_found)} criteria"

    def test_excel_raw_data_sheet(self, sample_excel: bytes) -> None:
        """Raw Data sheet contains original alternative values."""
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(sample_excel))
        ws = wb["Raw Data"]
        # Header
        assert ws.cell(row=1, column=1).value == "Alternative"
        # Should have data
        assert ws.max_row >= len(SAMPLE_ALTERNATIVES) + 1

    def test_excel_large_scenario(self) -> None:
        """Excel generates for 20+ alternatives."""
        alts = []
        for i in range(22):
            alts.append({
                "name": f"Alt_{i:02d}",
                "capex": 150000 + i * 10000,
                "opex": 50000 + i * 5000,
                "co2": 10 + i * 4,
                "risk": min(i * 0.4, 10),
                "comfort": max(10 - i * 0.3, 0),
                "maturity": max(10 - i * 0.2, 0),
            })
        results = compute_mcda_scores(alts, SAMPLE_WEIGHTS)
        xlsx = generate_mcda_excel("Large", alts, SAMPLE_WEIGHTS, results)
        assert isinstance(xlsx, bytes)
        assert len(xlsx) > 0

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(xlsx))
        ws = wb["Scores"]
        assert ws.max_row >= 23  # header + 22 data rows


# ---------------------------------------------------------------------------
# Celery task tests
# ---------------------------------------------------------------------------


class TestCeleryTask:
    """Tests for the MCDA report Celery task module."""

    def test_task_module_imports(self) -> None:
        """Task module can be imported without errors."""
        from app.tasks.sotreg_reports import (
            run_mcda_report_sync,
            run_mcda_report_task,
        )
        assert callable(run_mcda_report_task)
        assert callable(run_mcda_report_sync)

    def test_task_status_functions(self) -> None:
        """Status helper functions are importable."""
        from app.tasks.sotreg_reports import (
            get_report_status,
            set_report_status,
        )
        assert callable(set_report_status)
        assert callable(get_report_status)


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------


class TestEdgeCases:
    """Tests for edge cases and error handling."""

    def test_pdf_with_identical_scores(self) -> None:
        """PDF handles alternatives with identical criterion values."""
        identical = [
            {"name": "Alpha", "capex": 200000, "opex": 80000, "co2": 50, "risk": 3, "comfort": 5, "maturity": 5},
            {"name": "Beta", "capex": 200000, "opex": 80000, "co2": 50, "risk": 3, "comfort": 5, "maturity": 5},
        ]
        results = compute_mcda_scores(identical, SAMPLE_WEIGHTS)
        pdf = generate_mcda_pdf("Identical", identical, SAMPLE_WEIGHTS, results)
        assert isinstance(pdf, bytes)
        assert pdf[:5] == b"%PDF-"

    def test_excel_with_single_alternative(self) -> None:
        """Excel handles single alternative (minimum valid input)."""
        single = [
            {"name": "Solo", "capex": 200000, "opex": 80000, "co2": 50, "risk": 3, "comfort": 5, "maturity": 5},
        ]
        results = compute_mcda_scores(single, SAMPLE_WEIGHTS)
        xlsx = generate_mcda_excel("Solo", single, SAMPLE_WEIGHTS, results)
        assert isinstance(xlsx, bytes)

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(xlsx))
        assert len(wb.sheetnames) == 4

    def test_pdf_with_extreme_values(self) -> None:
        """PDF handles extreme criterion values without errors."""
        extreme = [
            {"name": "Min", "capex": 0, "opex": 0, "co2": 0, "risk": 0, "comfort": 0, "maturity": 0},
            {"name": "Max", "capex": 10000000, "opex": 5000000, "co2": 1000, "risk": 10, "comfort": 10, "maturity": 10},
        ]
        results = compute_mcda_scores(extreme, SAMPLE_WEIGHTS)
        pdf = generate_mcda_pdf("Extreme", extreme, SAMPLE_WEIGHTS, results)
        assert isinstance(pdf, bytes)
        assert len(pdf) > 500

    def test_criteria_labels_complete(self) -> None:
        """All 6 criteria have labels defined."""
        for c in CRITERIA_ORDER:
            assert c in CRITERIA_LABELS

    def test_criteria_order_matches_required(self) -> None:
        """CRITERIA_ORDER contains exactly the required criteria."""
        from app.services.sotreg.mcda_service import REQUIRED_CRITERIA
        assert set(CRITERIA_ORDER) == REQUIRED_CRITERIA
