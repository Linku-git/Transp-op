from __future__ import annotations

import csv
import io
import json
import uuid
from datetime import date, datetime, timezone
from types import SimpleNamespace
from typing import Any

import pytest

from app.services.export_engine import ExportEngine


# ---------------------------------------------------------------------------
# Helpers: build mock objects that behave like ORM instances
# ---------------------------------------------------------------------------


def _ns(**kwargs: Any) -> SimpleNamespace:
    """Create a SimpleNamespace with the given attributes."""
    return SimpleNamespace(**kwargs)


def _build_export_context() -> dict[str, Any]:
    """Build a complete mock context for the ExportEngine."""
    site_id = uuid.uuid4()
    opt_id = uuid.uuid4()
    cluster1_id = uuid.uuid4()
    cluster2_id = uuid.uuid4()
    route1_id = uuid.uuid4()
    route2_id = uuid.uuid4()
    vehicle1_id = uuid.uuid4()
    vehicle2_id = uuid.uuid4()
    tenant_id = uuid.uuid4()

    # 5 employees, 2 PMR
    emp_ids = [uuid.uuid4() for _ in range(5)]
    employees = {
        emp_ids[0]: _ns(
            id=emp_ids[0],
            matricule="EMP001",
            first_name="Ahmed",
            last_name="Benali",
            address="12 Rue Hassan II, Casablanca",
            city="Casablanca",
            lat=33.5731,
            lng=-7.5898,
            is_pmr=True,
            phone="+212600000001",
            department="Logistique",
            shift_time="Matin",
            site_id=site_id,
        ),
        emp_ids[1]: _ns(
            id=emp_ids[1],
            matricule="EMP002",
            first_name="Fatima",
            last_name="Zahra",
            address="45 Bd Anfa, Casablanca",
            city="Casablanca",
            lat=33.5850,
            lng=-7.6100,
            is_pmr=False,
            phone="+212600000002",
            department="RH",
            shift_time="Matin",
            site_id=site_id,
        ),
        emp_ids[2]: _ns(
            id=emp_ids[2],
            matricule="EMP003",
            first_name="Youssef",
            last_name="Alaoui",
            address="78 Ave FAR, Casablanca",
            city="Casablanca",
            lat=33.5900,
            lng=-7.6200,
            is_pmr=False,
            phone="+212600000003",
            department="IT",
            shift_time="Apres-midi",
            site_id=site_id,
        ),
        emp_ids[3]: _ns(
            id=emp_ids[3],
            matricule="EMP004",
            first_name="Sara",
            last_name="Idrissi",
            address="3 Rue Moulay Ismail, Casablanca",
            city="Casablanca",
            lat=33.5650,
            lng=-7.5800,
            is_pmr=True,
            phone="+212600000004",
            department="Finance",
            shift_time="Matin",
            site_id=site_id,
        ),
        emp_ids[4]: _ns(
            id=emp_ids[4],
            matricule="EMP005",
            first_name="Omar",
            last_name="Tazi",
            address="22 Rue de Fes, Casablanca",
            city="Casablanca",
            lat=33.5780,
            lng=-7.5950,
            is_pmr=False,
            phone="+212600000005",
            department="Logistique",
            shift_time="Matin",
            site_id=site_id,
        ),
    }

    # 2 vehicles
    vehicles = {
        vehicle1_id: _ns(
            id=vehicle1_id,
            type="Minibus",
            brand_model="Mercedes Sprinter",
            capacity=15,
            is_pmr_accessible=True,
            motorization="diesel",
            zfe_compliant=False,
            condition="Bon",
        ),
        vehicle2_id: _ns(
            id=vehicle2_id,
            type="Midibus",
            brand_model="Iveco Daily",
            capacity=22,
            is_pmr_accessible=False,
            motorization="hybrid",
            zfe_compliant=True,
            condition="Bon",
        ),
    }

    # 2 clusters
    clusters = [
        _ns(
            id=cluster1_id,
            optimization_id=opt_id,
            site_id=site_id,
            centroid_lat=33.5790,
            centroid_lng=-7.5950,
            employee_count=3,
            pmr_count=1,
            security_risk_level="low",
            employee_ids=[emp_ids[0], emp_ids[1], emp_ids[4]],
        ),
        _ns(
            id=cluster2_id,
            optimization_id=opt_id,
            site_id=site_id,
            centroid_lat=33.5775,
            centroid_lng=-7.6000,
            employee_count=2,
            pmr_count=1,
            security_risk_level="medium",
            employee_ids=[emp_ids[2], emp_ids[3]],
        ),
    ]

    # 2 routes with flat UUID-string stops (Format A — as produced by the pipeline)
    routes = [
        _ns(
            id=route1_id,
            optimization_id=opt_id,
            vehicle_id=vehicle1_id,
            site_id=site_id,
            ordered_stops=[str(emp_ids[0]), str(emp_ids[1]), str(emp_ids[4])],
            total_distance_km=12.5,
            total_time_minutes=25.0,
            polyline=None,
            rti_compliance_pct=95.0,
        ),
        _ns(
            id=route2_id,
            optimization_id=opt_id,
            vehicle_id=vehicle2_id,
            site_id=site_id,
            ordered_stops=[str(emp_ids[2]), str(emp_ids[3])],
            total_distance_km=8.3,
            total_time_minutes=18.0,
            polyline=None,
            rti_compliance_pct=88.5,
        ),
    ]

    # Site
    site = _ns(
        id=site_id,
        code="CASA-01",
        name="Casablanca HQ",
        address="Zone Industrielle, Ain Sebaa",
        city="Casablanca",
        lat=33.6000,
        lng=-7.5500,
    )

    # Optimization
    optimization = _ns(
        id=opt_id,
        tenant_id=tenant_id,
        site_id=site_id,
        condition_type="normal",
        status="completed",
        params={
            "algorithm": "dbscan",
            "eps_meters": 500.0,
            "min_samples": 2,
        },
        metrics={
            "total_employees": 5,
            "employees_assigned": 5,
            "total_clusters": 2,
            "total_vehicles_used": 2,
            "avg_occupancy_rate": 33.3,
            "total_distance_km": 20.8,
            "total_duration_minutes": 43.0,
            "estimated_fuel_cost_mad": 150.0,
            "co2_estimate_kg": 12.5,
        },
        target_date=date(2026, 4, 5),
        created_at=datetime(2026, 4, 2, 10, 0, 0, tzinfo=timezone.utc),
        completed_at=datetime(2026, 4, 2, 10, 2, 30, tzinfo=timezone.utc),
        site=site,
        clusters=clusters,
        routes=routes,
    )

    return {
        "optimization": optimization,
        "clusters": clusters,
        "routes": routes,
        "employees": employees,
        "vehicles": vehicles,
        "site": site,
    }


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestExportPdf:
    """test_export_pdf — Generates valid PDF file."""

    def test_generates_valid_pdf_bytes(self) -> None:
        context = _build_export_context()
        engine = ExportEngine(context)
        result = engine.generate_pdf()

        assert isinstance(result, bytes)
        assert len(result) > 0
        # PDF magic header
        assert result[:5] == b"%PDF-"

    def test_pdf_contains_multiple_pages(self) -> None:
        """PDF should have cover page + one page per route."""
        context = _build_export_context()
        engine = ExportEngine(context)
        result = engine.generate_pdf()
        # A multi-page PDF will contain multiple /Page entries
        assert result.count(b"/Type /Page") >= 2


class TestExportExcel:
    """test_export_excel — Generates valid XLSX with correct sheets."""

    def test_generates_valid_xlsx(self) -> None:
        from openpyxl import load_workbook

        context = _build_export_context()
        engine = ExportEngine(context)
        result = engine.generate_excel()

        assert isinstance(result, bytes)
        assert len(result) > 0

        wb = load_workbook(io.BytesIO(result))
        sheet_names = wb.sheetnames

        assert "Resume" in sheet_names
        assert "CASA-01" in sheet_names

    def test_resume_sheet_has_metrics(self) -> None:
        from openpyxl import load_workbook

        context = _build_export_context()
        engine = ExportEngine(context)
        result = engine.generate_excel()

        wb = load_workbook(io.BytesIO(result))
        ws = wb["Resume"]

        # Collect all cell values to check for metrics
        all_values = []
        for row in ws.iter_rows(values_only=True):
            all_values.extend([str(v) for v in row if v is not None])

        text = " ".join(all_values)
        assert "Employes total" in text
        assert "5" in text  # total_employees

    def test_site_sheet_has_employees(self) -> None:
        from openpyxl import load_workbook

        context = _build_export_context()
        engine = ExportEngine(context)
        result = engine.generate_excel()

        wb = load_workbook(io.BytesIO(result))
        ws = wb["CASA-01"]

        all_values = []
        for row in ws.iter_rows(values_only=True):
            all_values.extend([str(v) for v in row if v is not None])

        text = " ".join(all_values)
        assert "EMP001" in text
        assert "CLUSTERS" in text
        assert "ROUTES" in text
        assert "EMPLOYES" in text


class TestExportCsvStops:
    """test_export_csv_stops — CSV has correct columns and data."""

    def test_csv_has_correct_headers(self) -> None:
        context = _build_export_context()
        engine = ExportEngine(context)
        result = engine.generate_csv_stops()

        reader = csv.reader(io.StringIO(result))
        headers = next(reader)

        assert "route_number" in headers
        assert "stop_order" in headers
        assert "matricule" in headers
        assert "is_pmr" in headers
        assert "eta_seconds" in headers

    def test_csv_has_data_rows(self) -> None:
        context = _build_export_context()
        engine = ExportEngine(context)
        result = engine.generate_csv_stops()

        reader = csv.reader(io.StringIO(result))
        rows = list(reader)

        # Header + 5 stops (3 in route 1 + 2 in route 2)
        assert len(rows) == 6

    def test_csv_stop_order_sequential(self) -> None:
        context = _build_export_context()
        engine = ExportEngine(context)
        result = engine.generate_csv_stops()

        reader = csv.DictReader(io.StringIO(result))
        rows = list(reader)

        # Route 1 has stops 1,2,3
        route1_stops = [r for r in rows if r["route_number"] == "1"]
        assert [r["stop_order"] for r in route1_stops] == ["1", "2", "3"]


class TestExportCsvEmployees:
    """test_export_csv_employees — CSV has employee assignments."""

    def test_csv_has_correct_headers(self) -> None:
        context = _build_export_context()
        engine = ExportEngine(context)
        result = engine.generate_csv_employees()

        reader = csv.reader(io.StringIO(result))
        headers = next(reader)

        assert "matricule" in headers
        assert "is_pmr" in headers
        assert "cluster_id" in headers
        assert "route_number" in headers
        assert "vehicle_type" in headers

    def test_csv_has_all_employees(self) -> None:
        context = _build_export_context()
        engine = ExportEngine(context)
        result = engine.generate_csv_employees()

        reader = csv.reader(io.StringIO(result))
        rows = list(reader)

        # Header + 5 employees
        assert len(rows) == 6


class TestExportGeojson:
    """test_export_geojson — Valid GeoJSON FeatureCollection."""

    def test_valid_feature_collection(self) -> None:
        context = _build_export_context()
        engine = ExportEngine(context)
        result = engine.generate_geojson()

        assert isinstance(result, dict)
        assert result["type"] == "FeatureCollection"
        assert isinstance(result["features"], list)
        assert len(result["features"]) > 0

    def test_contains_all_feature_types(self) -> None:
        context = _build_export_context()
        engine = ExportEngine(context)
        result = engine.generate_geojson()

        feature_types = {f["properties"]["feature_type"] for f in result["features"]}
        assert "stop" in feature_types
        assert "cluster_centroid" in feature_types
        # Routes constructed from stops (no polyline) need >=2 coords
        # Route 1 has 3 stops, route 2 has 2 stops — both qualify
        assert "route" in feature_types

    def test_route_features_have_linestring(self) -> None:
        context = _build_export_context()
        engine = ExportEngine(context)
        result = engine.generate_geojson()

        route_features = [
            f for f in result["features"] if f["properties"]["feature_type"] == "route"
        ]
        for rf in route_features:
            assert rf["geometry"]["type"] == "LineString"
            coords = rf["geometry"]["coordinates"]
            assert len(coords) >= 2
            # GeoJSON: [lng, lat] — lng should be negative for Morocco
            for coord in coords:
                assert -180 <= coord[0] <= 180  # lng
                assert -90 <= coord[1] <= 90  # lat

    def test_stop_features_have_point(self) -> None:
        context = _build_export_context()
        engine = ExportEngine(context)
        result = engine.generate_geojson()

        stop_features = [
            f for f in result["features"] if f["properties"]["feature_type"] == "stop"
        ]
        assert len(stop_features) == 5  # 5 employees total
        for sf in stop_features:
            assert sf["geometry"]["type"] == "Point"

    def test_cluster_centroids_present(self) -> None:
        context = _build_export_context()
        engine = ExportEngine(context)
        result = engine.generate_geojson()

        cluster_features = [
            f
            for f in result["features"]
            if f["properties"]["feature_type"] == "cluster_centroid"
        ]
        assert len(cluster_features) == 2


class TestExportPmrIndicators:
    """test_export_with_pmr_indicators — PMR flags present in exports."""

    def test_csv_stops_pmr_present(self) -> None:
        context = _build_export_context()
        engine = ExportEngine(context)
        result = engine.generate_csv_stops()

        reader = csv.DictReader(io.StringIO(result))
        rows = list(reader)

        pmr_values = [r["is_pmr"] for r in rows]
        assert "True" in pmr_values
        assert "False" in pmr_values

    def test_csv_employees_pmr_present(self) -> None:
        context = _build_export_context()
        engine = ExportEngine(context)
        result = engine.generate_csv_employees()

        reader = csv.DictReader(io.StringIO(result))
        rows = list(reader)

        pmr_values = [r["is_pmr"] for r in rows]
        assert "True" in pmr_values
        assert "False" in pmr_values

    def test_geojson_stops_pmr_property(self) -> None:
        context = _build_export_context()
        engine = ExportEngine(context)
        result = engine.generate_geojson()

        stop_features = [
            f for f in result["features"] if f["properties"]["feature_type"] == "stop"
        ]
        pmr_values = [f["properties"]["is_pmr"] for f in stop_features]
        assert True in pmr_values
        assert False in pmr_values

    def test_excel_pmr_text_present(self) -> None:
        from openpyxl import load_workbook

        context = _build_export_context()
        engine = ExportEngine(context)
        result = engine.generate_excel()

        wb = load_workbook(io.BytesIO(result))
        ws = wb["CASA-01"]

        all_values = []
        for row in ws.iter_rows(values_only=True):
            all_values.extend([str(v) for v in row if v is not None])

        text = " ".join(all_values)
        assert "PMR" in text

    def test_pdf_generates_with_pmr_employees(self) -> None:
        """Verify PDF generates successfully with PMR employees in data.

        PDF content streams are compressed, so we verify the PDF is valid
        and has pages for routes containing PMR employees.
        """
        context = _build_export_context()
        engine = ExportEngine(context)
        result = engine.generate_pdf()

        assert isinstance(result, bytes)
        assert result[:5] == b"%PDF-"
        # 3 pages: cover + 2 routes (both contain PMR employees)
        assert result.count(b"/Type /Page") >= 3


class TestDecodePolyline:
    """Test the polyline decoder helper."""

    def test_decode_simple_polyline(self) -> None:
        # Encoded polyline for a simple 2-point line
        # This is the Google encoding for [(38.5, -120.2), (40.7, -120.95), (43.252, -126.453)]
        encoded = "_p~iF~ps|U_ulLnnqC_mqNvxq`@"
        coords = ExportEngine._decode_polyline(encoded)

        assert len(coords) == 3
        assert abs(coords[0][0] - 38.5) < 0.01
        assert abs(coords[0][1] - (-120.2)) < 0.01

    def test_decode_empty_polyline(self) -> None:
        coords = ExportEngine._decode_polyline("")
        assert coords == []
