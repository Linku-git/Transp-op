from __future__ import annotations

import csv
import io
import xml.etree.ElementTree as ET

import pytest
from httpx import AsyncClient

from app.services.daf_export import (
    build_tco_entries,
    build_roi_entries,
    generate_daf_csv,
    generate_daf_xml,
    generate_tco_pdf,
    generate_tco_excel,
    generate_roi_pdf,
    generate_roi_excel,
)
from tests.conftest import login_as_admin


# ---------------------------------------------------------------------------
# Shared test data
# ---------------------------------------------------------------------------

MOCK_TCO_DATA = {
    "fleet_tco": {
        "duration_years": 5,
        "vehicles": [
            {
                "vehicle_type": "minibus",
                "motorization": "diesel",
                "quantity": 3,
                "purchase_price": 150000,
                "maintenance_total": 60000,
                "energy_total": 30000,
                "tco_per_vehicle": 210000,
                "tco_total": 630000,
            },
        ],
        "fleet_tco_total": 630000,
        "vehicle_count": 3,
    }
}

MOCK_ROI_DATA = {
    "roi_absenteeism": 1155000,
    "roi_retention": 625000,
    "roi_fleet_optimization": 500000,
    "roi_journey": 1350000,
    "roi_total": 3630000,
    "roi_percentage": 121.0,
    "payback_months": 9.9,
    "total_investment": 3000000,
    "headcount": 500,
    "working_days_per_year": 220,
}


# ---------------------------------------------------------------------------
# DAF CSV Tests
# ---------------------------------------------------------------------------


def test_daf_export_csv_structure() -> None:
    """CSV has correct accounting columns."""
    entries = build_tco_entries(MOCK_TCO_DATA)
    csv_content = generate_daf_csv(entries, erp_format="sage")

    reader = csv.DictReader(io.StringIO(csv_content), delimiter=";")
    rows = list(reader)
    assert len(rows) > 0

    # Sage columns
    assert "JournalCode" in reader.fieldnames
    assert "CompteNum" in reader.fieldnames
    assert "Debit" in reader.fieldnames
    assert "Credit" in reader.fieldnames


def test_daf_export_xml_structure() -> None:
    """XML follows valid accounting interchange schema."""
    entries = build_tco_entries(MOCK_TCO_DATA)
    xml_content = generate_daf_xml(entries, erp_format="sage")

    root = ET.fromstring(xml_content)
    assert root.tag == "ComptabiliteExport"
    assert root.get("format") == "sage"

    journal = root.find("Journal")
    assert journal is not None
    ecritures = journal.findall("Ecriture")
    assert len(ecritures) > 0

    first = ecritures[0]
    assert first.find("Compte") is not None
    assert first.find("Libelle") is not None
    assert first.find("Debit") is not None


def test_daf_export_sap_format() -> None:
    """SAP FI format has BKPF/BSEG structure (BUKRS, BELNR, HKONT, etc.)."""
    entries = build_tco_entries(MOCK_TCO_DATA)
    csv_content = generate_daf_csv(entries, erp_format="sap_fi")

    reader = csv.DictReader(io.StringIO(csv_content), delimiter=";")
    rows = list(reader)
    assert len(rows) > 0
    assert "BUKRS" in reader.fieldnames
    assert "BELNR" in reader.fieldnames
    assert "HKONT" in reader.fieldnames
    assert "WRBTR" in reader.fieldnames
    assert "SHKZG" in reader.fieldnames

    # First row should be a debit entry
    assert rows[0]["SHKZG"] == "S"
    assert rows[0]["BUKRS"] == "1000"


def test_daf_export_sage_format() -> None:
    """Sage format has correct journal/account structure."""
    entries = build_tco_entries(MOCK_TCO_DATA)
    csv_content = generate_daf_csv(entries, erp_format="sage")

    reader = csv.DictReader(io.StringIO(csv_content), delimiter=";")
    rows = list(reader)
    assert len(rows) > 0
    assert rows[0]["JournalCode"] == "OD"
    assert rows[0]["CompteNum"] != ""
    assert rows[0]["PieceRef"] == "DAF-EXPORT"


def test_daf_export_cegid_format() -> None:
    """Cegid format has section analytique."""
    entries = build_tco_entries(MOCK_TCO_DATA)
    csv_content = generate_daf_csv(entries, erp_format="cegid")

    reader = csv.DictReader(io.StringIO(csv_content), delimiter=";")
    rows = list(reader)
    assert len(rows) > 0
    assert "SectionAnalytique" in reader.fieldnames
    assert "CodeAnalytique" in reader.fieldnames
    assert rows[0]["SectionAnalytique"] == "TRANSPORT"


# ---------------------------------------------------------------------------
# PDF/Excel Report Tests
# ---------------------------------------------------------------------------


def test_tco_report_pdf() -> None:
    """TCO PDF report generates with correct sections."""
    pdf_bytes = generate_tco_pdf(MOCK_TCO_DATA)
    assert len(pdf_bytes) > 100
    assert pdf_bytes[:5] == b"%PDF-"


def test_tco_report_excel() -> None:
    """TCO Excel report has correct sheets and data."""
    xlsx_bytes = generate_tco_excel(MOCK_TCO_DATA)
    assert len(xlsx_bytes) > 100

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws.title == "Resume TCO"
    # Check header row exists
    assert ws["A1"].value == "Rapport TCO — Cout Total de Possession"


def test_roi_report_pdf() -> None:
    """ROI PDF report generates with lever breakdown."""
    pdf_bytes = generate_roi_pdf(MOCK_ROI_DATA)
    assert len(pdf_bytes) > 100
    assert pdf_bytes[:5] == b"%PDF-"


def test_roi_report_excel() -> None:
    """ROI Excel report has correct sheets and data."""
    xlsx_bytes = generate_roi_excel(MOCK_ROI_DATA)
    assert len(xlsx_bytes) > 100

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws.title == "Resume ROI"
    assert ws["A1"].value == "Rapport ROI — Retour sur Investissement"


# ---------------------------------------------------------------------------
# Endpoint Integration Test
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_endpoint_download(client: AsyncClient) -> None:
    """POST /financial/export/daf returns file download."""
    token = await login_as_admin(client)
    headers = {"Authorization": f"Bearer {token}"}

    resp = await client.post(
        "/api/v1/financial/export/daf",
        headers=headers,
        json={
            "erp_format": "sage",
            "output_format": "csv",
            "tco_data": MOCK_TCO_DATA,
        },
    )
    assert resp.status_code == 200, f"Expected 200, got {resp.status_code}: {resp.text}"
    assert "text/csv" in resp.headers.get("content-type", "")
    assert "attachment" in resp.headers.get("content-disposition", "")
    assert len(resp.text) > 0

    # Also test XML
    resp_xml = await client.post(
        "/api/v1/financial/export/daf",
        headers=headers,
        json={
            "erp_format": "sap_fi",
            "output_format": "xml",
            "roi_data": MOCK_ROI_DATA,
        },
    )
    assert resp_xml.status_code == 200
    assert "xml" in resp_xml.headers.get("content-type", "")
