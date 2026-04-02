from __future__ import annotations

import csv
import io
import logging
import uuid
from datetime import datetime
from decimal import Decimal
from typing import Any

from fastapi import HTTPException, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.employee import Employee
from app.models.optimization import Cluster, Optimization, Route
from app.models.vehicle import Vehicle

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------


async def load_optimization_context(
    optimization_id: uuid.UUID,
    tenant_id: uuid.UUID,
    db: AsyncSession,
) -> dict[str, Any]:
    """Load the full optimization context needed for export.

    Returns a dict with keys: optimization, clusters, routes, employees,
    vehicles, site.
    """
    stmt = (
        select(Optimization)
        .options(
            selectinload(Optimization.clusters),
            selectinload(Optimization.routes),
        )
        .where(
            Optimization.id == optimization_id,
            Optimization.tenant_id == tenant_id,
        )
    )
    result = await db.execute(stmt)
    optimization = result.scalar_one_or_none()

    if optimization is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Optimization not found",
        )

    if optimization.status != "completed":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot export optimization with status '{optimization.status}'",
        )

    clusters = list(optimization.clusters)
    routes = list(optimization.routes)

    # Collect all employee UUIDs from clusters and route stops
    employee_ids: set[uuid.UUID] = set()
    for cluster in clusters:
        for eid in cluster.employee_ids or []:
            if isinstance(eid, uuid.UUID):
                employee_ids.add(eid)
            else:
                employee_ids.add(uuid.UUID(str(eid)))

    for route in routes:
        for stop in route.ordered_stops or []:
            if isinstance(stop, str):
                employee_ids.add(uuid.UUID(stop))
            elif isinstance(stop, dict) and stop.get("employee_id"):
                eid = stop["employee_id"]
                employee_ids.add(
                    eid if isinstance(eid, uuid.UUID) else uuid.UUID(str(eid))
                )

    # Batch-load employees
    employees_map: dict[uuid.UUID, Employee] = {}
    if employee_ids:
        emp_stmt = select(Employee).where(Employee.id.in_(list(employee_ids)))
        emp_result = await db.execute(emp_stmt)
        for emp in emp_result.scalars().all():
            employees_map[emp.id] = emp

    # Batch-load vehicles
    vehicle_ids: set[uuid.UUID] = set()
    for route in routes:
        if route.vehicle_id is not None:
            vehicle_ids.add(route.vehicle_id)

    vehicles_map: dict[uuid.UUID, Vehicle] = {}
    if vehicle_ids:
        veh_stmt = select(Vehicle).where(Vehicle.id.in_(list(vehicle_ids)))
        veh_result = await db.execute(veh_stmt)
        for veh in veh_result.scalars().all():
            vehicles_map[veh.id] = veh

    return {
        "optimization": optimization,
        "clusters": clusters,
        "routes": routes,
        "employees": employees_map,
        "vehicles": vehicles_map,
        "site": optimization.site,
    }


# ---------------------------------------------------------------------------
# Export engine
# ---------------------------------------------------------------------------


class ExportEngine:
    """Generates export files from an optimization context."""

    def __init__(self, context: dict[str, Any]) -> None:
        self.optimization: Optimization = context["optimization"]
        self.clusters: list[Cluster] = context["clusters"]
        self.routes: list[Route] = context["routes"]
        self.employees: dict[uuid.UUID, Employee] = context["employees"]
        self.vehicles: dict[uuid.UUID, Vehicle] = context["vehicles"]
        self.site = context["site"]

    # -- Helpers ----------------------------------------------------------

    def _parse_ordered_stops(self, route: Route) -> list[dict[str, Any]]:
        """Normalise ordered_stops from either flat UUID list or rich dicts."""
        stops = route.ordered_stops or []
        if not stops:
            return []

        # Format A: flat list of UUID strings
        if isinstance(stops[0], str):
            normalised: list[dict[str, Any]] = []
            for eid_str in stops:
                eid = uuid.UUID(eid_str)
                emp = self.employees.get(eid)
                normalised.append(
                    {
                        "employee_id": eid_str,
                        "lat": emp.lat if emp else None,
                        "lng": emp.lng if emp else None,
                        "is_pickup": True,
                        "eta_seconds": 0,
                        "cumulative_distance_meters": 0,
                    }
                )
            return normalised

        # Format B: rich dicts
        return stops

    def _build_employee_assignments(
        self,
    ) -> dict[uuid.UUID, dict[str, Any]]:
        """Map each employee to their cluster and route assignment."""
        assignments: dict[uuid.UUID, dict[str, Any]] = {}

        for idx, cluster in enumerate(self.clusters):
            for eid in cluster.employee_ids or []:
                eid_uuid = eid if isinstance(eid, uuid.UUID) else uuid.UUID(str(eid))
                assignments.setdefault(eid_uuid, {})["cluster_index"] = idx + 1
                assignments[eid_uuid]["cluster_id"] = str(cluster.id)

        for route_idx, route in enumerate(self.routes):
            for stop in self._parse_ordered_stops(route):
                eid_str = stop.get("employee_id")
                if eid_str:
                    eid_uuid = (
                        eid_str
                        if isinstance(eid_str, uuid.UUID)
                        else uuid.UUID(str(eid_str))
                    )
                    assignments.setdefault(eid_uuid, {})["route_number"] = route_idx + 1
                    assignments[eid_uuid]["vehicle_id"] = route.vehicle_id

        return assignments

    @staticmethod
    def _decode_polyline(encoded: str) -> list[tuple[float, float]]:
        """Decode Google encoded polyline to list of (lat, lng) tuples."""
        coords: list[tuple[float, float]] = []
        index = 0
        lat = 0
        lng = 0
        length = len(encoded)

        while index < length:
            # Decode latitude
            shift = 0
            result = 0
            while True:
                b = ord(encoded[index]) - 63
                index += 1
                result |= (b & 0x1F) << shift
                shift += 5
                if b < 0x20:
                    break
            lat += (~(result >> 1) if (result & 1) else (result >> 1))

            # Decode longitude
            shift = 0
            result = 0
            while True:
                b = ord(encoded[index]) - 63
                index += 1
                result |= (b & 0x1F) << shift
                shift += 5
                if b < 0x20:
                    break
            lng += (~(result >> 1) if (result & 1) else (result >> 1))

            coords.append((lat / 1e5, lng / 1e5))

        return coords

    @staticmethod
    def _decimal_to_float(val: Decimal | float | None) -> float | None:
        if val is None:
            return None
        return float(val)

    def _format_eta(self, eta_seconds: float | int | None) -> str:
        """Format ETA seconds as HH:MM display string."""
        if eta_seconds is None or eta_seconds == 0:
            return ""
        total_min = int(float(eta_seconds) / 60)
        hours = total_min // 60
        minutes = total_min % 60
        if hours > 0:
            return f"{hours}h{minutes:02d}"
        return f"{minutes}min"

    # -- CSV stop order ---------------------------------------------------

    def generate_csv_stops(self) -> str:
        """Generate CSV of ordered stops across all routes."""
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(
            [
                "route_number",
                "stop_order",
                "employee_id",
                "matricule",
                "first_name",
                "last_name",
                "address",
                "city",
                "lat",
                "lng",
                "is_pmr",
                "is_pickup",
                "eta_seconds",
                "cumulative_distance_m",
            ]
        )

        for route_idx, route in enumerate(self.routes):
            stops = self._parse_ordered_stops(route)
            for stop_idx, stop in enumerate(stops):
                eid_str = stop.get("employee_id", "")
                emp = None
                if eid_str:
                    eid = (
                        eid_str
                        if isinstance(eid_str, uuid.UUID)
                        else uuid.UUID(str(eid_str))
                    )
                    emp = self.employees.get(eid)

                writer.writerow(
                    [
                        route_idx + 1,
                        stop_idx + 1,
                        str(eid_str),
                        emp.matricule if emp else "",
                        emp.first_name if emp else "",
                        emp.last_name if emp else "",
                        emp.address if emp else "",
                        emp.city if emp else "",
                        stop.get("lat", ""),
                        stop.get("lng", ""),
                        emp.is_pmr if emp else False,
                        stop.get("is_pickup", True),
                        stop.get("eta_seconds", 0),
                        stop.get("cumulative_distance_meters", 0),
                    ]
                )

        return buffer.getvalue()

    # -- CSV employee assignments -----------------------------------------

    def generate_csv_employees(self) -> str:
        """Generate CSV of employee assignments (cluster + route)."""
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(
            [
                "employee_id",
                "matricule",
                "first_name",
                "last_name",
                "department",
                "site_code",
                "site_name",
                "address",
                "city",
                "lat",
                "lng",
                "is_pmr",
                "phone",
                "shift_time",
                "cluster_id",
                "route_number",
                "vehicle_type",
            ]
        )

        assignments = self._build_employee_assignments()

        for eid, emp in self.employees.items():
            asgn = assignments.get(eid, {})
            vehicle_type = ""
            vid = asgn.get("vehicle_id")
            if vid:
                veh = self.vehicles.get(vid)
                vehicle_type = veh.type if veh else ""

            writer.writerow(
                [
                    str(eid),
                    emp.matricule,
                    emp.first_name,
                    emp.last_name,
                    emp.department or "",
                    self.site.code if self.site else "",
                    self.site.name if self.site else "",
                    emp.address or "",
                    emp.city or "",
                    emp.lat or "",
                    emp.lng or "",
                    emp.is_pmr,
                    emp.phone or "",
                    emp.shift_time or "",
                    asgn.get("cluster_id", ""),
                    asgn.get("route_number", ""),
                    vehicle_type,
                ]
            )

        return buffer.getvalue()

    # -- GeoJSON ----------------------------------------------------------

    def generate_geojson(self) -> dict[str, Any]:
        """Generate GeoJSON FeatureCollection with routes, stops, clusters."""
        features: list[dict[str, Any]] = []

        # Route LineStrings
        for route_idx, route in enumerate(self.routes):
            coords: list[list[float]] = []

            if route.polyline:
                decoded = self._decode_polyline(route.polyline)
                coords = [[lng, lat] for lat, lng in decoded]
            else:
                # Construct from stops
                stops = self._parse_ordered_stops(route)
                for stop in stops:
                    lat = stop.get("lat")
                    lng = stop.get("lng")
                    if lat is not None and lng is not None:
                        coords.append([float(lng), float(lat)])

            if len(coords) >= 2:
                vehicle = self.vehicles.get(route.vehicle_id) if route.vehicle_id else None
                features.append(
                    {
                        "type": "Feature",
                        "geometry": {
                            "type": "LineString",
                            "coordinates": coords,
                        },
                        "properties": {
                            "feature_type": "route",
                            "route_id": str(route.id),
                            "route_number": route_idx + 1,
                            "vehicle_id": str(route.vehicle_id) if route.vehicle_id else None,
                            "vehicle_type": vehicle.type if vehicle else None,
                            "total_distance_km": self._decimal_to_float(route.total_distance_km),
                            "total_time_minutes": self._decimal_to_float(route.total_time_minutes),
                            "rti_compliance_pct": self._decimal_to_float(route.rti_compliance_pct),
                        },
                    }
                )

        # Stop Points
        for route_idx, route in enumerate(self.routes):
            stops = self._parse_ordered_stops(route)
            for stop_idx, stop in enumerate(stops):
                lat = stop.get("lat")
                lng = stop.get("lng")
                if lat is None or lng is None:
                    continue

                eid_str = stop.get("employee_id", "")
                emp = None
                if eid_str:
                    eid = (
                        eid_str
                        if isinstance(eid_str, uuid.UUID)
                        else uuid.UUID(str(eid_str))
                    )
                    emp = self.employees.get(eid)

                features.append(
                    {
                        "type": "Feature",
                        "geometry": {
                            "type": "Point",
                            "coordinates": [float(lng), float(lat)],
                        },
                        "properties": {
                            "feature_type": "stop",
                            "employee_id": str(eid_str),
                            "matricule": emp.matricule if emp else None,
                            "name": f"{emp.first_name} {emp.last_name}" if emp else None,
                            "is_pmr": emp.is_pmr if emp else False,
                            "is_pickup": stop.get("is_pickup", True),
                            "stop_order": stop_idx + 1,
                            "route_number": route_idx + 1,
                            "eta_seconds": stop.get("eta_seconds", 0),
                        },
                    }
                )

        # Cluster centroid Points
        for idx, cluster in enumerate(self.clusters):
            features.append(
                {
                    "type": "Feature",
                    "geometry": {
                        "type": "Point",
                        "coordinates": [cluster.centroid_lng, cluster.centroid_lat],
                    },
                    "properties": {
                        "feature_type": "cluster_centroid",
                        "cluster_id": str(cluster.id),
                        "cluster_number": idx + 1,
                        "employee_count": cluster.employee_count,
                        "pmr_count": cluster.pmr_count,
                    },
                }
            )

        return {
            "type": "FeatureCollection",
            "features": features,
        }

    # -- Excel ------------------------------------------------------------

    def generate_excel(self) -> bytes:
        """Generate multi-sheet Excel workbook."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        pmr_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

        # -- Resume sheet -------------------------------------------------
        ws = wb.active
        ws.title = "Resume"

        ws.append(["Rapport d'Optimisation"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])

        site_name = self.site.name if self.site else "N/A"
        site_code = self.site.code if self.site else "N/A"

        info_rows = [
            ["Identifiant", str(self.optimization.id)],
            ["Site", f"{site_name} ({site_code})"],
            ["Condition", self.optimization.condition_type],
            ["Statut", self.optimization.status],
            ["Date cible", str(self.optimization.target_date or "N/A")],
            ["Cree le", str(self.optimization.created_at)],
            ["Termine le", str(self.optimization.completed_at or "N/A")],
        ]
        for row in info_rows:
            ws.append(row)

        ws.append([])
        ws.append(["Metriques"])
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=12)

        metrics = self.optimization.metrics or {}
        metric_labels = {
            "total_employees": "Employes total",
            "employees_assigned": "Employes affectes",
            "total_clusters": "Clusters total",
            "total_vehicles_used": "Vehicules utilises",
            "avg_occupancy_rate": "Taux d'occupation moyen (%)",
            "total_distance_km": "Distance totale (km)",
            "total_duration_minutes": "Duree totale (min)",
            "estimated_fuel_cost_mad": "Cout carburant estime (MAD)",
            "co2_estimate_kg": "CO2 estime (kg)",
        }
        for key, label in metric_labels.items():
            val = metrics.get(key, "N/A")
            ws.append([label, val])

        # Adjust column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 40

        # -- Per-site sheet -----------------------------------------------
        sheet_name = site_code[:31] if site_code else "Site"
        ws_site = wb.create_sheet(title=sheet_name)

        # Clusters section
        ws_site.append(["CLUSTERS"])
        ws_site[f"A{ws_site.max_row}"].font = Font(bold=True, size=12)

        cluster_headers = [
            "Cluster #",
            "Centroid Lat",
            "Centroid Lng",
            "Employes",
            "PMR",
            "Risque",
        ]
        ws_site.append(cluster_headers)
        for col_idx, _ in enumerate(cluster_headers, 1):
            cell = ws_site.cell(row=ws_site.max_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        for idx, cluster in enumerate(self.clusters):
            ws_site.append(
                [
                    idx + 1,
                    cluster.centroid_lat,
                    cluster.centroid_lng,
                    cluster.employee_count,
                    cluster.pmr_count,
                    cluster.security_risk_level,
                ]
            )

        ws_site.append([])

        # Routes section
        ws_site.append(["ROUTES"])
        ws_site[f"A{ws_site.max_row}"].font = Font(bold=True, size=12)

        route_headers = [
            "Route #",
            "Type vehicule",
            "Modele",
            "Capacite",
            "PMR accessible",
            "Distance (km)",
            "Duree (min)",
            "Arrets",
            "RTI (%)",
        ]
        ws_site.append(route_headers)
        for col_idx, _ in enumerate(route_headers, 1):
            cell = ws_site.cell(row=ws_site.max_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        for idx, route in enumerate(self.routes):
            vehicle = self.vehicles.get(route.vehicle_id) if route.vehicle_id else None
            stops = self._parse_ordered_stops(route)
            ws_site.append(
                [
                    idx + 1,
                    vehicle.type if vehicle else "N/A",
                    vehicle.brand_model if vehicle else "N/A",
                    vehicle.capacity if vehicle else "N/A",
                    "Oui" if (vehicle and vehicle.is_pmr_accessible) else "Non",
                    self._decimal_to_float(route.total_distance_km) or "N/A",
                    self._decimal_to_float(route.total_time_minutes) or "N/A",
                    len(stops),
                    self._decimal_to_float(route.rti_compliance_pct) or "N/A",
                ]
            )

        ws_site.append([])

        # Employees section
        ws_site.append(["EMPLOYES"])
        ws_site[f"A{ws_site.max_row}"].font = Font(bold=True, size=12)

        emp_headers = [
            "Matricule",
            "Nom",
            "Prenom",
            "Departement",
            "Adresse",
            "PMR",
            "Cluster #",
            "Route #",
        ]
        ws_site.append(emp_headers)
        for col_idx, _ in enumerate(emp_headers, 1):
            cell = ws_site.cell(row=ws_site.max_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        assignments = self._build_employee_assignments()
        for eid, emp in self.employees.items():
            asgn = assignments.get(eid, {})
            row_data = [
                emp.matricule,
                emp.last_name,
                emp.first_name,
                emp.department or "",
                emp.address or "",
                "PMR" if emp.is_pmr else "",
                asgn.get("cluster_index", ""),
                asgn.get("route_number", ""),
            ]
            ws_site.append(row_data)

            # Highlight PMR rows
            if emp.is_pmr:
                for col_idx in range(1, len(row_data) + 1):
                    ws_site.cell(row=ws_site.max_row, column=col_idx).fill = pmr_fill

        # Adjust column widths for site sheet
        for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
            ws_site.column_dimensions[col_letter].width = 18

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # -- PDF --------------------------------------------------------------

    def generate_pdf(self) -> bytes:
        """Generate PDF driver sheets using reportlab."""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            PageBreak,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=1.5 * cm,
            rightMargin=1.5 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            spaceAfter=12,
        )
        subtitle_style = ParagraphStyle(
            "CustomSubtitle",
            parent=styles["Heading2"],
            fontSize=12,
            spaceAfter=8,
        )
        normal_style = styles["Normal"]

        story: list[Any] = []
        site_name = self.site.name if self.site else "N/A"

        # -- Cover page ---------------------------------------------------
        story.append(Paragraph(f"Feuilles de Route - {site_name}", title_style))
        story.append(
            Paragraph(
                f"Optimisation: {str(self.optimization.id)[:8]}... | "
                f"Date: {self.optimization.target_date or self.optimization.created_at.date()}",
                subtitle_style,
            )
        )
        story.append(Spacer(1, 0.5 * cm))

        # Summary metrics table
        metrics = self.optimization.metrics or {}
        summary_data = [
            ["Metrique", "Valeur"],
            ["Employes total", str(metrics.get("total_employees", "N/A"))],
            ["Employes affectes", str(metrics.get("employees_assigned", "N/A"))],
            ["Clusters", str(metrics.get("total_clusters", "N/A"))],
            ["Vehicules utilises", str(metrics.get("total_vehicles_used", "N/A"))],
            [
                "Taux d'occupation",
                f"{metrics.get('avg_occupancy_rate', 'N/A')}%",
            ],
            [
                "Distance totale",
                f"{metrics.get('total_distance_km', 'N/A')} km",
            ],
            ["CO2 estime", f"{metrics.get('co2_estimate_kg', 'N/A')} kg"],
        ]

        summary_table = Table(summary_data, colWidths=[10 * cm, 7 * cm])
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0058be")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.append(summary_table)

        # -- Per-route driver sheets --------------------------------------
        for route_idx, route in enumerate(self.routes):
            story.append(PageBreak())

            vehicle = self.vehicles.get(route.vehicle_id) if route.vehicle_id else None
            v_type = vehicle.type if vehicle else "N/A"
            v_model = vehicle.brand_model if vehicle else ""
            v_cap = vehicle.capacity if vehicle else "N/A"
            v_pmr = "Oui" if (vehicle and vehicle.is_pmr_accessible) else "Non"
            dist = self._decimal_to_float(route.total_distance_km)
            dur = self._decimal_to_float(route.total_time_minutes)

            story.append(
                Paragraph(
                    f"Route {route_idx + 1} - {v_type} {v_model}",
                    title_style,
                )
            )
            story.append(
                Paragraph(
                    f"Capacite: {v_cap} | PMR: {v_pmr} | "
                    f"Distance: {dist or 'N/A'} km | "
                    f"Duree: {dur or 'N/A'} min",
                    subtitle_style,
                )
            )
            story.append(Spacer(1, 0.3 * cm))

            # Stops table
            stops = self._parse_ordered_stops(route)
            table_data = [["#", "Matricule", "Nom complet", "Adresse", "PMR", "ETA"]]
            pmr_rows: list[int] = []

            for stop_idx, stop in enumerate(stops):
                eid_str = stop.get("employee_id", "")
                emp = None
                if eid_str:
                    eid = (
                        eid_str
                        if isinstance(eid_str, uuid.UUID)
                        else uuid.UUID(str(eid_str))
                    )
                    emp = self.employees.get(eid)

                full_name = f"{emp.first_name} {emp.last_name}" if emp else "N/A"
                address = emp.address or "" if emp else ""
                is_pmr = emp.is_pmr if emp else False
                eta_str = self._format_eta(stop.get("eta_seconds"))

                row = [
                    stop_idx + 1,
                    emp.matricule if emp else "",
                    full_name,
                    address[:40],  # Truncate long addresses
                    "PMR" if is_pmr else "",
                    eta_str,
                ]
                table_data.append(row)

                if is_pmr:
                    pmr_rows.append(stop_idx + 1)  # 1-based (header is row 0)

            if len(table_data) > 1:
                stops_table = Table(
                    table_data,
                    colWidths=[1.2 * cm, 2.5 * cm, 4.5 * cm, 5.5 * cm, 1.5 * cm, 2 * cm],
                )
                style_commands: list[Any] = [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0058be")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]

                # Highlight PMR rows
                for pmr_row in pmr_rows:
                    style_commands.append(
                        ("BACKGROUND", (0, pmr_row), (-1, pmr_row), colors.HexColor("#E3F2FD"))
                    )
                    style_commands.append(
                        ("FONTNAME", (4, pmr_row), (4, pmr_row), "Helvetica-Bold")
                    )
                    style_commands.append(
                        ("TEXTCOLOR", (4, pmr_row), (4, pmr_row), colors.HexColor("#D32F2F"))
                    )

                stops_table.setStyle(TableStyle(style_commands))
                story.append(stops_table)
            else:
                story.append(Paragraph("Aucun arret pour cette route.", normal_style))

        doc.build(story)
        return buffer.getvalue()
