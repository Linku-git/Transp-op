from __future__ import annotations

import csv
import io
import logging
from datetime import date, datetime
from decimal import Decimal
from typing import Any
from xml.etree.ElementTree import Element, SubElement, tostring

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# ERP Format Templates
# ---------------------------------------------------------------------------

ERP_FORMATS = {
    "sap_fi": {
        "label": "SAP FI",
        "columns": [
            "BUKRS", "BELNR", "GJAHR", "BLDAT", "BUDAT", "BLART",
            "BSCHL", "HKONT", "WRBTR", "SHKZG", "SGTXT",
        ],
        "journal_code": "FI",
        "doc_type": "SA",
    },
    "sage": {
        "label": "Sage",
        "columns": [
            "JournalCode", "JournalLib", "EcritureNum", "EcritureDate",
            "CompteNum", "CompteLib", "PieceRef", "PieceDate",
            "EcritureLib", "Debit", "Credit",
        ],
        "journal_code": "OD",
    },
    "cegid": {
        "label": "Cegid",
        "columns": [
            "CodeJournal", "DateEcriture", "NumeroCompte", "LibelleEcriture",
            "Debit", "Credit", "SectionAnalytique", "CodeAnalytique",
        ],
        "journal_code": "OD",
    },
}

ACCOUNT_CODES = {
    "transport_cost": {"code": "625100", "label": "Frais de transport du personnel"},
    "vehicle_purchase": {"code": "218200", "label": "Materiel de transport"},
    "maintenance": {"code": "615200", "label": "Entretien et reparations"},
    "fuel": {"code": "606100", "label": "Carburants"},
    "insurance": {"code": "616000", "label": "Primes d'assurance"},
    "depreciation": {"code": "681120", "label": "Dotations amort. materiel transport"},
    "roi_absenteeism": {"code": "791000", "label": "Transfert charges - absenteisme"},
    "roi_retention": {"code": "791100", "label": "Transfert charges - retention"},
    "roi_fleet": {"code": "791200", "label": "Transfert charges - flotte"},
    "roi_journey": {"code": "791300", "label": "Transfert charges - trajet"},
}


def _fmt(val: float | Decimal) -> str:
    """Format number as 2-decimal string."""
    return f"{float(val):.2f}"


# ---------------------------------------------------------------------------
# CSV Export
# ---------------------------------------------------------------------------


def generate_daf_csv(
    entries: list[dict[str, Any]],
    erp_format: str = "sage",
) -> str:
    """Generate ERP-compatible CSV with accounting entries."""
    fmt = ERP_FORMATS.get(erp_format, ERP_FORMATS["sage"])
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fmt["columns"], delimiter=";")
    writer.writeheader()

    for entry in entries:
        row = _format_entry_for_erp(entry, erp_format, fmt)
        writer.writerow(row)

    return output.getvalue()


def _format_entry_for_erp(
    entry: dict[str, Any],
    erp_format: str,
    fmt: dict[str, Any],
) -> dict[str, str]:
    """Map a generic accounting entry to ERP-specific column format."""
    today = date.today().strftime("%Y%m%d")
    today_slash = date.today().strftime("%d/%m/%Y")

    if erp_format == "sap_fi":
        return {
            "BUKRS": "1000",
            "BELNR": entry.get("doc_number", "0000000001"),
            "GJAHR": str(date.today().year),
            "BLDAT": today,
            "BUDAT": today,
            "BLART": fmt["doc_type"],
            "BSCHL": "40" if entry.get("debit", 0) > 0 else "50",
            "HKONT": entry["account_code"],
            "WRBTR": _fmt(entry.get("debit", 0) or entry.get("credit", 0)),
            "SHKZG": "S" if entry.get("debit", 0) > 0 else "H",
            "SGTXT": entry["label"],
        }
    elif erp_format == "cegid":
        return {
            "CodeJournal": fmt["journal_code"],
            "DateEcriture": today_slash,
            "NumeroCompte": entry["account_code"],
            "LibelleEcriture": entry["label"],
            "Debit": _fmt(entry.get("debit", 0)),
            "Credit": _fmt(entry.get("credit", 0)),
            "SectionAnalytique": entry.get("section", "TRANSPORT"),
            "CodeAnalytique": entry.get("analytic_code", "MOBILITE"),
        }
    else:  # sage
        return {
            "JournalCode": fmt["journal_code"],
            "JournalLib": "Operations Diverses",
            "EcritureNum": entry.get("doc_number", "00001"),
            "EcritureDate": today_slash,
            "CompteNum": entry["account_code"],
            "CompteLib": entry["label"],
            "PieceRef": entry.get("piece_ref", "DAF-EXPORT"),
            "PieceDate": today_slash,
            "EcritureLib": entry.get("description", entry["label"]),
            "Debit": _fmt(entry.get("debit", 0)),
            "Credit": _fmt(entry.get("credit", 0)),
        }


# ---------------------------------------------------------------------------
# XML Export
# ---------------------------------------------------------------------------


def generate_daf_xml(
    entries: list[dict[str, Any]],
    erp_format: str = "sage",
) -> str:
    """Generate accounting XML export."""
    root = Element("ComptabiliteExport")
    root.set("format", erp_format)
    root.set("date", datetime.now().isoformat())

    journal = SubElement(root, "Journal")
    journal.set("code", ERP_FORMATS.get(erp_format, ERP_FORMATS["sage"])["journal_code"])

    for i, entry in enumerate(entries, 1):
        ecriture = SubElement(journal, "Ecriture")
        ecriture.set("numero", str(i))

        SubElement(ecriture, "Compte").text = entry["account_code"]
        SubElement(ecriture, "Libelle").text = entry["label"]
        SubElement(ecriture, "Debit").text = _fmt(entry.get("debit", 0))
        SubElement(ecriture, "Credit").text = _fmt(entry.get("credit", 0))
        SubElement(ecriture, "Date").text = date.today().isoformat()

        if erp_format == "cegid":
            SubElement(ecriture, "SectionAnalytique").text = entry.get("section", "TRANSPORT")

    return tostring(root, encoding="unicode", xml_declaration=True)


# ---------------------------------------------------------------------------
# Build accounting entries from financial data
# ---------------------------------------------------------------------------


def build_tco_entries(tco_data: dict[str, Any]) -> list[dict[str, Any]]:
    """Build accounting entries from TCO calculation results."""
    entries: list[dict[str, Any]] = []
    fleet = tco_data.get("fleet_tco", tco_data)

    for vehicle in fleet.get("vehicles", []):
        qty = vehicle.get("quantity", 1)
        entries.append({
            "account_code": ACCOUNT_CODES["vehicle_purchase"]["code"],
            "label": f"{ACCOUNT_CODES['vehicle_purchase']['label']} - {vehicle['vehicle_type']}",
            "debit": vehicle.get("purchase_price", 0) * qty,
            "credit": 0,
        })
        entries.append({
            "account_code": ACCOUNT_CODES["maintenance"]["code"],
            "label": f"{ACCOUNT_CODES['maintenance']['label']} - {vehicle['vehicle_type']}",
            "debit": vehicle.get("maintenance_total", 0),
            "credit": 0,
        })
        entries.append({
            "account_code": ACCOUNT_CODES["fuel"]["code"],
            "label": f"{ACCOUNT_CODES['fuel']['label']} - {vehicle['vehicle_type']}",
            "debit": vehicle.get("energy_total", 0),
            "credit": 0,
        })

    return entries


def build_roi_entries(roi_data: dict[str, Any]) -> list[dict[str, Any]]:
    """Build accounting entries from ROI calculation results."""
    entries: list[dict[str, Any]] = []

    lever_map = {
        "roi_absenteeism": "roi_absenteeism",
        "roi_retention": "roi_retention",
        "roi_fleet_optimization": "roi_fleet",
        "roi_journey": "roi_journey",
    }

    for key, acct_key in lever_map.items():
        value = roi_data.get(key, 0)
        if value and value > 0:
            entries.append({
                "account_code": ACCOUNT_CODES[acct_key]["code"],
                "label": ACCOUNT_CODES[acct_key]["label"],
                "debit": 0,
                "credit": value,
            })

    return entries


# ---------------------------------------------------------------------------
# PDF Report Generators
# ---------------------------------------------------------------------------


def generate_tco_pdf(tco_data: dict[str, Any]) -> bytes:
    """Generate TCO report as PDF."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    elements: list[Any] = []

    # Title
    elements.append(Paragraph("Rapport TCO — Cout Total de Possession", styles["Title"]))
    elements.append(Spacer(1, 10 * mm))

    fleet = tco_data.get("fleet_tco", tco_data)

    # Summary
    elements.append(Paragraph("Resume", styles["Heading2"]))
    summary_data = [
        ["Metrique", "Valeur"],
        ["TCO Total Flotte", f"{fleet.get('fleet_tco_total', 0):,.2f} MAD"],
        ["Nombre de vehicules", str(fleet.get("vehicle_count", 0))],
        ["Duree (annees)", str(fleet.get("duration_years", 0))],
    ]
    t = Table(summary_data, colWidths=[200, 200])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0058be")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 8 * mm))

    # Vehicle breakdown
    if fleet.get("vehicles"):
        elements.append(Paragraph("Detail par vehicule", styles["Heading2"]))
        headers = ["Type", "Motorisation", "Qte", "TCO/vehicule", "TCO Total"]
        rows = [headers]
        for v in fleet["vehicles"]:
            rows.append([
                v.get("vehicle_type", ""),
                v.get("motorization", ""),
                str(v.get("quantity", 1)),
                f"{v.get('tco_per_vehicle', 0):,.2f}",
                f"{v.get('tco_total', 0):,.2f}",
            ])
        t2 = Table(rows, colWidths=[100, 90, 40, 100, 100])
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0058be")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(t2)

    doc.build(elements)
    return buffer.getvalue()


def generate_roi_pdf(roi_data: dict[str, Any]) -> bytes:
    """Generate ROI report as PDF."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    elements: list[Any] = []

    elements.append(Paragraph("Rapport ROI — Retour sur Investissement", styles["Title"]))
    elements.append(Spacer(1, 10 * mm))

    # Summary
    elements.append(Paragraph("Resume", styles["Heading2"]))
    summary = [
        ["Metrique", "Valeur"],
        ["ROI Total", f"{roi_data.get('roi_total', 0):,.2f} MAD"],
        ["ROI %", f"{roi_data.get('roi_percentage', 0):.1f}%"],
        ["Payback", f"{roi_data.get('payback_months', 'N/A')} mois"],
        ["Investissement", f"{roi_data.get('total_investment', 0):,.2f} MAD"],
        ["Effectif", str(roi_data.get("headcount", 0))],
    ]
    t = Table(summary, colWidths=[200, 200])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0058be")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 8 * mm))

    # Lever breakdown
    elements.append(Paragraph("Detail des leviers ROI", styles["Heading2"]))
    levers = [
        ["Levier", "Montant (MAD)"],
        ["Absenteisme", f"{roi_data.get('roi_absenteeism', 0):,.2f}"],
        ["Retention", f"{roi_data.get('roi_retention', 0):,.2f}"],
        ["Optimisation flotte", f"{roi_data.get('roi_fleet_optimization', 0):,.2f}"],
        ["Productivite trajet", f"{roi_data.get('roi_journey', 0):,.2f}"],
        ["TOTAL", f"{roi_data.get('roi_total', 0):,.2f}"],
    ]
    t2 = Table(levers, colWidths=[200, 200])
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0058be")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    elements.append(t2)

    doc.build(elements)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Excel Report Generators
# ---------------------------------------------------------------------------


HEADER_FILL = PatternFill(start_color="0058be", end_color="0058be", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def generate_tco_excel(tco_data: dict[str, Any]) -> bytes:
    """Generate TCO report as Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resume TCO"

    fleet = tco_data.get("fleet_tco", tco_data)

    # Summary
    ws.append(["Rapport TCO — Cout Total de Possession"])
    ws.merge_cells("A1:E1")
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["Metrique", "Valeur"])
    for cell in ws[3]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.append(["TCO Total Flotte (MAD)", fleet.get("fleet_tco_total", 0)])
    ws.append(["Nombre de vehicules", fleet.get("vehicle_count", 0)])
    ws.append(["Duree (annees)", fleet.get("duration_years", 0)])
    ws.append([])

    # Vehicle detail
    headers = ["Type", "Motorisation", "Quantite", "Achat", "Maintenance", "Energie", "TCO/vehicule", "TCO Total"]
    ws.append(headers)
    row_num = ws.max_row
    for cell in ws[row_num]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for v in fleet.get("vehicles", []):
        ws.append([
            v.get("vehicle_type", ""),
            v.get("motorization", ""),
            v.get("quantity", 1),
            v.get("purchase_price", 0),
            v.get("maintenance_total", 0),
            v.get("energy_total", 0),
            v.get("tco_per_vehicle", 0),
            v.get("tco_total", 0),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_roi_excel(roi_data: dict[str, Any]) -> bytes:
    """Generate ROI report as Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resume ROI"

    ws.append(["Rapport ROI — Retour sur Investissement"])
    ws.merge_cells("A1:C1")
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["Metrique", "Valeur"])
    for cell in ws[3]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.append(["ROI Total (MAD)", roi_data.get("roi_total", 0)])
    ws.append(["ROI (%)", roi_data.get("roi_percentage", 0)])
    ws.append(["Payback (mois)", roi_data.get("payback_months", "N/A")])
    ws.append(["Investissement (MAD)", roi_data.get("total_investment", 0)])
    ws.append(["Effectif", roi_data.get("headcount", 0)])
    ws.append([])

    # Lever detail
    ws.append(["Levier", "Montant (MAD)"])
    row_num = ws.max_row
    for cell in ws[row_num]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    ws.append(["Absenteisme", roi_data.get("roi_absenteeism", 0)])
    ws.append(["Retention", roi_data.get("roi_retention", 0)])
    ws.append(["Optimisation flotte", roi_data.get("roi_fleet_optimization", 0)])
    ws.append(["Productivite trajet", roi_data.get("roi_journey", 0)])
    ws.append(["TOTAL", roi_data.get("roi_total", 0)])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
