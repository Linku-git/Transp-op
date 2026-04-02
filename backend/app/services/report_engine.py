from __future__ import annotations

import csv
import io
import logging
import uuid
from typing import Any

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.employee import Employee
from app.models.modal import EmployeeModal
from app.models.optimization import Optimization
from app.models.site import Site
from app.models.vehicle import Vehicle

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helper: PDF table builder (reportlab)
# ---------------------------------------------------------------------------


def _build_pdf(title: str, headers: list[str], rows: list[list[Any]]) -> bytes:
    """Build a simple single-table PDF report using reportlab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontSize=18,
        spaceAfter=12,
    )

    story: list[Any] = []
    story.append(Paragraph(title, title_style))
    story.append(Spacer(1, 0.5 * cm))

    table_data = [headers] + rows

    # Compute column widths proportionally
    num_cols = len(headers)
    available_width = landscape(A4)[0] - 3 * cm
    col_width = available_width / num_cols
    col_widths = [col_width] * num_cols

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0058be")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#F0F4F8")],
                ),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(table)

    doc.build(story)
    return buffer.getvalue()


def _build_excel(
    title: str, headers: list[str], rows: list[list[Any]]
) -> bytes:
    """Build a single-sheet Excel workbook with styled header."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(
        start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"
    )

    # Title row
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    # Headers
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    for row in rows:
        ws.append(row)

    # Auto-size columns
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[
            ws.cell(row=3, column=col_idx).column_letter
        ].width = max(len(str(header)) + 4, 14)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _build_csv(headers: list[str], rows: list[list[Any]]) -> bytes:
    """Build a CSV file as bytes."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8")


# ---------------------------------------------------------------------------
# Modal Analysis Report
# ---------------------------------------------------------------------------


async def generate_modal_analysis_report(
    tenant_id: uuid.UUID,
    db: AsyncSession,
    report_format: str = "pdf",
) -> bytes:
    """Generate modal analysis report with mode distribution and travel stats."""
    # Mode distribution
    mode_stmt = (
        select(
            EmployeeModal.primary_mode,
            func.count().label("count"),
            func.coalesce(func.avg(EmployeeModal.distance_km), 0).label("avg_distance"),
            func.coalesce(func.avg(EmployeeModal.travel_time_min), 0).label("avg_time"),
            func.coalesce(func.min(EmployeeModal.distance_km), 0).label("min_distance"),
            func.coalesce(func.max(EmployeeModal.distance_km), 0).label("max_distance"),
        )
        .join(Employee, EmployeeModal.employee_id == Employee.id)
        .where(Employee.tenant_id == tenant_id, Employee.active.is_(True))
        .group_by(EmployeeModal.primary_mode)
        .order_by(func.count().desc())
    )
    result = await db.execute(mode_stmt)
    mode_rows = result.all()

    title = "Rapport Analyse Modale"
    headers = [
        "Mode de transport",
        "Nombre",
        "Distance moy. (km)",
        "Temps moy. (min)",
        "Distance min (km)",
        "Distance max (km)",
    ]

    rows = [
        [
            r.primary_mode,
            r.count,
            round(float(r.avg_distance), 1),
            round(float(r.avg_time), 0),
            round(float(r.min_distance), 1),
            round(float(r.max_distance), 1),
        ]
        for r in mode_rows
    ]

    # Add total row
    total_count = sum(r.count for r in mode_rows)
    rows.append(["TOTAL", total_count, "", "", "", ""])

    if report_format == "xlsx":
        return _build_excel(title, headers, rows)
    return _build_pdf(title, headers, rows)


# ---------------------------------------------------------------------------
# Fleet Utilization Report
# ---------------------------------------------------------------------------


async def generate_fleet_utilization_report(
    tenant_id: uuid.UUID,
    db: AsyncSession,
    report_format: str = "pdf",
) -> bytes:
    """Generate fleet utilization report with vehicle breakdown."""
    # Vehicle summary by type, motorization, condition
    vehicle_stmt = (
        select(
            Vehicle.type,
            Vehicle.motorization,
            Vehicle.condition,
            func.count().label("count"),
            func.sum(Vehicle.capacity).label("total_capacity"),
            func.avg(Vehicle.capacity).label("avg_capacity"),
        )
        .where(Vehicle.tenant_id == tenant_id)
        .group_by(Vehicle.type, Vehicle.motorization, Vehicle.condition)
        .order_by(Vehicle.type, Vehicle.motorization)
    )
    result = await db.execute(vehicle_stmt)
    vehicle_rows = result.all()

    # Latest optimization metrics for utilization
    opt_stmt = (
        select(Optimization.metrics)
        .where(
            Optimization.tenant_id == tenant_id,
            Optimization.status == "completed",
        )
        .order_by(Optimization.completed_at.desc())
        .limit(1)
    )
    opt_result = await db.execute(opt_stmt)
    opt_metrics = opt_result.scalar_one_or_none() or {}

    title = "Rapport Utilisation Flotte"
    headers = [
        "Type",
        "Motorisation",
        "Condition",
        "Nombre",
        "Capacite totale",
        "Capacite moy.",
    ]

    rows = [
        [
            r.type,
            r.motorization or "N/A",
            r.condition,
            r.count,
            int(r.total_capacity) if r.total_capacity else 0,
            round(float(r.avg_capacity), 1) if r.avg_capacity else 0,
        ]
        for r in vehicle_rows
    ]

    # Add summary row from latest optimization
    total_vehicles = sum(r.count for r in vehicle_rows)
    total_capacity = sum(
        int(r.total_capacity) if r.total_capacity else 0 for r in vehicle_rows
    )
    vehicles_used = opt_metrics.get("total_vehicles_used", "N/A") if opt_metrics else "N/A"
    occupancy = opt_metrics.get("avg_occupancy_rate", "N/A") if opt_metrics else "N/A"

    rows.append([])
    rows.append(["TOTAL", "", "", total_vehicles, total_capacity, ""])
    rows.append(["Vehicules utilises (derniere optimisation)", "", "", vehicles_used, "", ""])
    rows.append(["Taux d'occupation moyen (%)", "", "", occupancy, "", ""])

    if report_format == "xlsx":
        return _build_excel(title, headers, rows)
    return _build_pdf(title, headers, rows)


# ---------------------------------------------------------------------------
# Volunteer Driver Report
# ---------------------------------------------------------------------------


async def generate_volunteer_driver_report(
    tenant_id: uuid.UUID,
    db: AsyncSession,
    report_format: str = "pdf",
) -> bytes:
    """Generate volunteer driver report with driver details."""
    driver_stmt = (
        select(
            Employee.matricule,
            Employee.first_name,
            Employee.last_name,
            Employee.department,
            Employee.shift_time,
            Employee.carpool_seats,
            Site.name.label("site_name"),
        )
        .join(Site, Employee.site_id == Site.id)
        .where(
            Employee.tenant_id == tenant_id,
            Employee.active.is_(True),
            Employee.volunteer_driver.is_(True),
        )
        .order_by(Site.name, Employee.last_name)
    )
    result = await db.execute(driver_stmt)
    driver_rows = result.all()

    title = "Rapport Conducteurs Volontaires"
    headers = [
        "Matricule",
        "Prenom",
        "Nom",
        "Departement",
        "Equipe/Poste",
        "Places covoiturage",
        "Site",
    ]

    rows = [
        [
            r.matricule,
            r.first_name,
            r.last_name,
            r.department or "N/A",
            r.shift_time or "N/A",
            r.carpool_seats,
            r.site_name,
        ]
        for r in driver_rows
    ]

    # Summary
    total_drivers = len(driver_rows)
    total_seats = sum(r.carpool_seats for r in driver_rows)
    rows.append([])
    rows.append(["TOTAL", "", "", "", "", total_seats, f"{total_drivers} conducteurs"])

    if report_format == "xlsx":
        return _build_excel(title, headers, rows)
    return _build_pdf(title, headers, rows)


# ---------------------------------------------------------------------------
# HR Mobility Report
# ---------------------------------------------------------------------------


async def generate_hr_mobility_report(
    tenant_id: uuid.UUID,
    db: AsyncSession,
    report_format: str = "pdf",
) -> bytes:
    """Generate HR mobility report using HR KPIs (coverage, shadow zones, etc.)."""
    from app.services.hr_analytics import (
        compute_mobility_coverage,
        compute_shadow_zones,
    )

    coverage = await compute_mobility_coverage(tenant_id, db)
    shadow = await compute_shadow_zones(tenant_id, db)

    # Build coverage-by-site table
    title = "Rapport Mobilite RH"
    headers = [
        "Site",
        "Total employes",
        "Employes couverts",
        "Couverture (%)",
    ]

    rows: list[list[Any]] = []
    for site_data in coverage.get("by_site", []):
        rows.append([
            site_data["name"],
            site_data["total"],
            site_data["covered"],
            site_data["pct"],
        ])

    # Global summary rows
    rows.append([])
    rows.append([
        "GLOBAL",
        coverage.get("total_employees", 0),
        coverage.get("covered_employees", 0),
        coverage.get("coverage_pct", 0),
    ])
    rows.append([])
    rows.append(["ZONES D'OMBRE", "", "", ""])
    rows.append([
        f"Employes en zone d'ombre: {shadow.get('shadow_zone_count', 0)}",
        f"Total actifs: {shadow.get('total_active_employees', 0)}",
        f"Taux: {shadow.get('shadow_zone_pct', 0)}%",
        f"Seuil: {shadow.get('threshold_km', 30)} km",
    ])

    if report_format == "xlsx":
        return _build_excel(title, headers, rows)
    return _build_pdf(title, headers, rows)
