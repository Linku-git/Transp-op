"""MCDA Comparison Report Generator — PDF & Excel exports.

Produces structured PDF reports with executive summary, comparison tables,
horizontal scoring bars, radar charts (6 axes), sensitivity tables, and
recommendation sections.  Also generates multi-sheet Excel workbooks with
conditional formatting.

Session 113 — CDC SOTREG v5.0 Module M7.
"""
from __future__ import annotations

import io
import logging
import math
import uuid
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.mcda_scenario import MCDAScenario
from app.services.sotreg.mcda_service import (
    CDC_DEFAULT_WEIGHTS,
    CRITERIA_DIRECTION,
    REQUIRED_CRITERIA,
    compute_mcda_scores,
    compute_sensitivity_analysis,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

CRITERIA_LABELS: dict[str, str] = {
    "capex": "CAPEX",
    "opex": "OPEX",
    "co2": "CO₂",
    "risk": "Risque",
    "comfort": "Confort",
    "maturity": "Maturité",
}

CRITERIA_ORDER: list[str] = ["capex", "opex", "co2", "risk", "comfort", "maturity"]

# Color palette for up to 10 alternatives
ALT_COLORS: list[str] = [
    "#0058be",  # Azure Blue (primary)
    "#e63946",  # Red
    "#2a9d8f",  # Teal
    "#e9c46a",  # Yellow
    "#f4a261",  # Orange
    "#264653",  # Dark Teal
    "#6a4c93",  # Purple
    "#1d3557",  # Navy
    "#457b9d",  # Steel Blue
    "#a8dadc",  # Light Teal
]


# ---------------------------------------------------------------------------
# Helpers — load scenario from DB
# ---------------------------------------------------------------------------


async def _load_scenario(
    scenario_id: uuid.UUID,
    db: AsyncSession,
) -> MCDAScenario | None:
    """Fetch an MCDAScenario by primary key."""
    stmt = select(MCDAScenario).where(MCDAScenario.id == scenario_id)
    result = await db.execute(stmt)
    return result.scalar_one_or_none()


# ---------------------------------------------------------------------------
# Radar chart drawing (reportlab canvas)
# ---------------------------------------------------------------------------


def _draw_radar_chart(
    canvas: Any,
    cx: float,
    cy: float,
    radius: float,
    alternatives: list[dict],
    criteria: list[str],
    colors: list[str],
) -> None:
    """Draw a radar/spider chart on a reportlab canvas.

    Args:
        canvas: ReportLab canvas object.
        cx, cy: Center coordinates.
        radius: Chart radius in points.
        alternatives: List of dicts with ``name`` and ``normalized_values``.
        criteria: Ordered list of criteria keys.
        colors: Color hex codes per alternative.
    """
    from reportlab.lib.colors import HexColor

    n = len(criteria)
    if n < 3:
        return

    angle_step = 2 * math.pi / n

    # Draw grid rings (1–5 scale)
    canvas.setStrokeColor(HexColor("#c2c6d6"))
    canvas.setLineWidth(0.3)
    for ring in range(1, 6):
        r = radius * ring / 5
        points = []
        for i in range(n):
            angle = -math.pi / 2 + i * angle_step
            px = cx + r * math.cos(angle)
            py = cy + r * math.sin(angle)
            points.append((px, py))
        path = canvas.beginPath()
        path.moveTo(*points[0])
        for pt in points[1:]:
            path.lineTo(*pt)
        path.close()
        canvas.drawPath(path, fill=0, stroke=1)

    # Draw axes
    canvas.setStrokeColor(HexColor("#424754"))
    canvas.setLineWidth(0.5)
    for i in range(n):
        angle = -math.pi / 2 + i * angle_step
        ex = cx + radius * math.cos(angle)
        ey = cy + radius * math.sin(angle)
        canvas.line(cx, cy, ex, ey)

    # Draw axis labels
    canvas.setFont("Helvetica-Bold", 8)
    canvas.setFillColor(HexColor("#191c1e"))
    label_margin = 14
    for i, criterion in enumerate(criteria):
        angle = -math.pi / 2 + i * angle_step
        lx = cx + (radius + label_margin) * math.cos(angle)
        ly = cy + (radius + label_margin) * math.sin(angle)
        label = CRITERIA_LABELS.get(criterion, criterion)
        canvas.drawCentredString(lx, ly - 3, label)

    # Draw alternative polygons
    for alt_idx, alt in enumerate(alternatives):
        color = HexColor(colors[alt_idx % len(colors)])
        nv = alt.get("normalized_values", {})

        points = []
        for i, criterion in enumerate(criteria):
            val = nv.get(criterion, 3.0)
            r = radius * val / 5.0
            angle = -math.pi / 2 + i * angle_step
            px = cx + r * math.cos(angle)
            py = cy + r * math.sin(angle)
            points.append((px, py))

        # Fill with transparency
        canvas.saveState()
        canvas.setFillColor(color, alpha=0.15)
        canvas.setStrokeColor(color, alpha=0.8)
        canvas.setLineWidth(1.5)
        path = canvas.beginPath()
        path.moveTo(*points[0])
        for pt in points[1:]:
            path.lineTo(*pt)
        path.close()
        canvas.drawPath(path, fill=1, stroke=1)
        canvas.restoreState()

        # Draw dots at vertices
        canvas.setFillColor(color)
        for px, py in points:
            canvas.circle(px, py, 2.5, fill=1, stroke=0)


# ---------------------------------------------------------------------------
# PDF generation
# ---------------------------------------------------------------------------


def generate_mcda_pdf(
    scenario_name: str,
    alternatives: list[dict],
    weights: dict[str, float],
    results: dict,
) -> bytes:
    """Generate a structured MCDA comparison PDF report.

    Sections:
      1. Executive summary
      2. Comparison table
      3. Horizontal scoring bars
      4. Radar chart (6 axes, overlapping polygons)
      5. Sensitivity table
      6. Recommendation

    Args:
        scenario_name: Name of the MCDA scenario.
        alternatives: Raw alternative dicts.
        weights: Weights used for scoring.
        results: Full results dict from ``compute_mcda_scores``.

    Returns:
        PDF bytes.
    """
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.colors import HexColor
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2.5 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        "MCDATitle",
        parent=styles["Heading1"],
        fontSize=20,
        textColor=HexColor("#0058be"),
        spaceAfter=6,
    )
    section_style = ParagraphStyle(
        "SectionHead",
        parent=styles["Heading2"],
        fontSize=13,
        textColor=HexColor("#191c1e"),
        spaceBefore=14,
        spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "MCDABody",
        parent=styles["Normal"],
        fontSize=9,
        leading=13,
        textColor=HexColor("#424754"),
    )
    small_style = ParagraphStyle(
        "MCDASmall",
        parent=styles["Normal"],
        fontSize=8,
        leading=11,
        textColor=HexColor("#424754"),
    )

    ranked = results.get("ranked_alternatives", [])
    best = results.get("best_alternative", "N/A")
    worst = results.get("worst_alternative", "N/A")

    # Compute sensitivity for the report
    sensitivity = compute_sensitivity_analysis(alternatives, weights)

    story: list[Any] = []

    # ===== TITLE =====
    story.append(Paragraph(f"Rapport MCDA — {scenario_name}", title_style))
    story.append(Spacer(1, 4 * mm))

    # ===== 1. EXECUTIVE SUMMARY =====
    story.append(Paragraph("1. Résumé Exécutif", section_style))
    n_alts = len(ranked)
    best_score = ranked[0]["score"] if ranked else 0
    worst_score = ranked[-1]["score"] if ranked else 0

    summary_text = (
        f"Cette analyse multicritère compare <b>{n_alts}</b> alternatives de transport "
        f"sur 6 critères pondérés (CAPEX, OPEX, CO₂, Risque, Confort, Maturité). "
        f"L'alternative <b>{best}</b> obtient le meilleur score global "
        f"({best_score:.2f}/5.00), tandis que <b>{worst}</b> se classe en "
        f"dernière position ({worst_score:.2f}/5.00)."
    )
    story.append(Paragraph(summary_text, body_style))
    story.append(Spacer(1, 2 * mm))

    stability = sensitivity.get("stability_score", 100)
    critical = sensitivity.get("critical_criteria", [])
    stability_text = (
        f"Score de stabilité : <b>{stability:.0f}%</b>. "
    )
    if critical:
        crit_labels = ", ".join(CRITERIA_LABELS.get(c, c) for c in critical)
        stability_text += f"Critères critiques : {crit_labels}."
    else:
        stability_text += "Aucun critère critique détecté — le classement est robuste."
    story.append(Paragraph(stability_text, body_style))
    story.append(Spacer(1, 4 * mm))

    # ===== 2. COMPARISON TABLE =====
    story.append(Paragraph("2. Tableau Comparatif", section_style))

    table_headers = ["Rang", "Alternative", "Score"]
    for c in CRITERIA_ORDER:
        table_headers.append(CRITERIA_LABELS.get(c, c))

    table_data = [table_headers]
    for alt in ranked:
        row = [
            str(alt["rank"]),
            alt["name"],
            f"{alt['score']:.2f}",
        ]
        nv = alt.get("normalized_values", {})
        for c in CRITERIA_ORDER:
            row.append(f"{nv.get(c, 0):.2f}")
        table_data.append(row)

    n_cols = len(table_headers)
    col_widths = [30, 80, 45] + [50] * 6
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), HexColor("#0058be")),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#c2c6d6")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, HexColor("#F0F4F8")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (2, 0), (-1, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            # Highlight best row
            ("BACKGROUND", (0, 1), (-1, 1), HexColor("#E8F4E8")),
        ])
    )
    story.append(table)
    story.append(Spacer(1, 4 * mm))

    # ===== 3. SCORING BARS =====
    story.append(Paragraph("3. Barres de Score", section_style))

    bar_data = []
    max_score = max((a["score"] for a in ranked), default=5)
    for alt in ranked:
        score = alt["score"]
        bar_width = int(score / max_score * 30) if max_score > 0 else 0
        bar_visual = "█" * bar_width + "░" * (30 - bar_width)
        bar_data.append([
            alt["name"],
            f"{score:.2f}",
            bar_visual,
        ])

    bar_table = Table(bar_data, colWidths=[90, 45, 300])
    bar_table.setStyle(
        TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("FONTNAME", (2, 0), (2, -1), "Courier"),
            ("FONTSIZE", (2, 0), (2, -1), 7),
            ("TEXTCOLOR", (2, 0), (2, -1), HexColor("#0058be")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ])
    )
    story.append(bar_table)
    story.append(Spacer(1, 4 * mm))

    # ===== 4. RADAR CHART =====
    story.append(Paragraph("4. Graphique Radar (6 axes)", section_style))

    # Draw radar on a separate drawing using reportlab canvas callback
    from reportlab.platypus import Flowable

    class RadarFlowable(Flowable):
        """Custom flowable that draws a radar chart."""

        def __init__(
            self,
            width: float,
            height: float,
            alts: list[dict],
            criteria: list[str],
            colors: list[str],
        ):
            super().__init__()
            self.width = width
            self.height = height
            self._alts = alts
            self._criteria = criteria
            self._colors = colors

        def draw(self) -> None:
            cx = self.width / 2
            cy = self.height / 2
            chart_radius = min(cx, cy) - 30
            _draw_radar_chart(
                self.canv, cx, cy, chart_radius,
                self._alts, self._criteria, self._colors,
            )
            # Draw legend below chart
            legend_y = 12
            for i, alt in enumerate(self._alts):
                color = HexColor(self._colors[i % len(self._colors)])
                lx = 30 + i * 110
                self.canv.setFillColor(color)
                self.canv.rect(lx, legend_y, 8, 8, fill=1, stroke=0)
                self.canv.setFillColor(HexColor("#191c1e"))
                self.canv.setFont("Helvetica", 7)
                self.canv.drawString(lx + 12, legend_y + 1, alt["name"][:18])

    radar = RadarFlowable(
        width=400, height=300,
        alts=ranked, criteria=CRITERIA_ORDER, colors=ALT_COLORS,
    )
    story.append(radar)
    story.append(Spacer(1, 4 * mm))

    # ===== 5. SENSITIVITY TABLE =====
    story.append(Paragraph("5. Analyse de Sensibilité", section_style))

    sens_headers = [
        "Critère", "Poids", "Poids +Δ", "Poids −Δ",
        "Classement modifié", "Critique",
    ]
    sens_data = [sens_headers]
    for sr in sensitivity.get("sensitivity_results", []):
        sens_data.append([
            CRITERIA_LABELS.get(sr["criterion"], sr["criterion"]),
            f"{sr['weight_original']:.2f}",
            f"{sr['weight_plus']:.2f}",
            f"{sr['weight_minus']:.2f}",
            "Oui" if sr["ranking_changed"] else "Non",
            "⚠ Critique" if sr["is_critical"] else "Stable",
        ])

    sens_table = Table(sens_data, colWidths=[70, 55, 55, 55, 90, 70])
    sens_style_cmds: list[Any] = [
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#0058be")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#c2c6d6")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, HexColor("#F0F4F8")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    # Highlight critical rows
    for row_idx, sr in enumerate(sensitivity.get("sensitivity_results", []), start=1):
        if sr["is_critical"]:
            sens_style_cmds.append(
                ("BACKGROUND", (0, row_idx), (-1, row_idx), HexColor("#FFF3E0"))
            )
    sens_table.setStyle(TableStyle(sens_style_cmds))
    story.append(sens_table)
    story.append(Spacer(1, 2 * mm))

    story.append(
        Paragraph(
            f"Score de stabilité global : <b>{stability:.0f}%</b> "
            f"(delta ±20%)",
            small_style,
        )
    )
    story.append(Spacer(1, 4 * mm))

    # ===== 6. RECOMMENDATION =====
    story.append(Paragraph("6. Recommandation", section_style))

    if stability >= 80:
        confidence = "élevée"
    elif stability >= 50:
        confidence = "modérée"
    else:
        confidence = "faible"

    rec_text = (
        f"Sur la base de l'analyse multicritère pondérée, l'alternative "
        f"<b>{best}</b> est recommandée avec un score de {best_score:.2f}/5.00. "
        f"La confiance dans ce classement est <b>{confidence}</b> "
        f"(stabilité {stability:.0f}%)."
    )
    story.append(Paragraph(rec_text, body_style))
    story.append(Spacer(1, 2 * mm))

    # Weights used
    weights_text = "Pondérations appliquées : " + ", ".join(
        f"{CRITERIA_LABELS.get(k, k)} = {v:.0%}" for k, v in sorted(weights.items())
    )
    story.append(Paragraph(weights_text, small_style))

    doc.build(story)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------


def generate_mcda_excel(
    scenario_name: str,
    alternatives: list[dict],
    weights: dict[str, float],
    results: dict,
) -> bytes:
    """Generate a multi-sheet MCDA comparison Excel workbook.

    Sheets:
      1. Summary — executive summary and top-level metrics
      2. Scores — full comparison table with normalized scores
      3. Sensitivity — weight perturbation analysis
      4. Raw Data — original alternative values

    Includes conditional formatting (green-yellow-red) on score cells.

    Args:
        scenario_name: Name of the MCDA scenario.
        alternatives: Raw alternative dicts.
        weights: Weights used.
        results: Full results from ``compute_mcda_scores``.

    Returns:
        Excel (.xlsx) bytes.
    """
    from openpyxl import Workbook
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    ranked = results.get("ranked_alternatives", [])
    best = results.get("best_alternative", "N/A")
    worst = results.get("worst_alternative", "N/A")

    # Style presets
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="0058BE", end_color="0058BE", fill_type="solid")
    title_font = Font(bold=True, size=14, color="0058BE")
    subtitle_font = Font(bold=True, size=11, color="191C1E")
    thin_border = Border(
        left=Side(style="thin", color="C2C6D6"),
        right=Side(style="thin", color="C2C6D6"),
        top=Side(style="thin", color="C2C6D6"),
        bottom=Side(style="thin", color="C2C6D6"),
    )

    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append([f"Rapport MCDA — {scenario_name}"])
    ws_summary["A1"].font = title_font
    ws_summary.append([])
    ws_summary.append(["Métrique", "Valeur"])
    ws_summary["A3"].font = subtitle_font
    ws_summary["B3"].font = subtitle_font

    summary_rows = [
        ("Nombre d'alternatives", len(ranked)),
        ("Meilleure alternative", best),
        ("Score meilleur", ranked[0]["score"] if ranked else "N/A"),
        ("Pire alternative", worst),
        ("Score pire", ranked[-1]["score"] if ranked else "N/A"),
    ]

    # Compute sensitivity for Excel
    sensitivity = compute_sensitivity_analysis(alternatives, weights)
    summary_rows.append(("Stabilité (%)", sensitivity.get("stability_score", 100)))
    critical = sensitivity.get("critical_criteria", [])
    summary_rows.append((
        "Critères critiques",
        ", ".join(CRITERIA_LABELS.get(c, c) for c in critical) if critical else "Aucun",
    ))

    for label, value in summary_rows:
        ws_summary.append([label, value])

    ws_summary.append([])
    ws_summary.append(["Pondérations"])
    ws_summary[f"A{ws_summary.max_row}"].font = subtitle_font
    for c in CRITERIA_ORDER:
        ws_summary.append([CRITERIA_LABELS.get(c, c), f"{weights.get(c, 0):.0%}"])

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 25

    # --- Sheet 2: Scores ---
    ws_scores = wb.create_sheet("Scores")
    score_headers = ["Rang", "Alternative", "Score Global"]
    for c in CRITERIA_ORDER:
        score_headers.append(CRITERIA_LABELS.get(c, c))

    ws_scores.append(score_headers)
    for col_idx in range(1, len(score_headers) + 1):
        cell = ws_scores.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for alt in ranked:
        nv = alt.get("normalized_values", {})
        row = [alt["rank"], alt["name"], alt["score"]]
        for c in CRITERIA_ORDER:
            row.append(nv.get(c, 0))
        ws_scores.append(row)

    # Conditional formatting on score columns (columns D-I, rows 2+)
    n_data_rows = len(ranked)
    if n_data_rows > 0:
        for col_idx in range(4, 10):  # Columns D through I (6 criteria)
            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}2:{col_letter}{n_data_rows + 1}"
            ws_scores.conditional_formatting.add(
                cell_range,
                ColorScaleRule(
                    start_type="num", start_value=1, start_color="F44336",
                    mid_type="num", mid_value=3, mid_color="FFC107",
                    end_type="num", end_value=5, end_color="4CAF50",
                ),
            )
        # Also format global score (column C)
        global_range = f"C2:C{n_data_rows + 1}"
        ws_scores.conditional_formatting.add(
            global_range,
            ColorScaleRule(
                start_type="min", start_color="F44336",
                mid_type="percentile", mid_value=50, mid_color="FFC107",
                end_type="max", end_color="4CAF50",
            ),
        )

    # Auto-size columns
    for col_idx in range(1, len(score_headers) + 1):
        ws_scores.column_dimensions[get_column_letter(col_idx)].width = max(
            len(str(score_headers[col_idx - 1])) + 4, 14
        )

    # Apply borders to data
    for row_idx in range(2, n_data_rows + 2):
        for col_idx in range(1, len(score_headers) + 1):
            ws_scores.cell(row=row_idx, column=col_idx).border = thin_border
            ws_scores.cell(row=row_idx, column=col_idx).alignment = Alignment(
                horizontal="center"
            )

    # --- Sheet 3: Sensitivity ---
    ws_sens = wb.create_sheet("Sensitivity")
    sens_headers = [
        "Critère", "Poids original", "Poids +Δ", "Poids −Δ",
        "Classement modifié", "Critique",
    ]
    ws_sens.append(sens_headers)
    for col_idx in range(1, len(sens_headers) + 1):
        cell = ws_sens.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    critical_fill = PatternFill(
        start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"
    )
    for row_idx, sr in enumerate(sensitivity.get("sensitivity_results", []), start=2):
        ws_sens.cell(row=row_idx, column=1, value=CRITERIA_LABELS.get(
            sr["criterion"], sr["criterion"]
        ))
        ws_sens.cell(row=row_idx, column=2, value=sr["weight_original"])
        ws_sens.cell(row=row_idx, column=3, value=sr["weight_plus"])
        ws_sens.cell(row=row_idx, column=4, value=sr["weight_minus"])
        ws_sens.cell(row=row_idx, column=5, value="Oui" if sr["ranking_changed"] else "Non")
        ws_sens.cell(row=row_idx, column=6, value="Critique" if sr["is_critical"] else "Stable")

        if sr["is_critical"]:
            for col_idx in range(1, 7):
                ws_sens.cell(row=row_idx, column=col_idx).fill = critical_fill

    ws_sens.append([])
    ws_sens.append(["Score de stabilité", f"{sensitivity.get('stability_score', 100):.0f}%"])

    for col_idx in range(1, len(sens_headers) + 1):
        ws_sens.column_dimensions[get_column_letter(col_idx)].width = 18

    # --- Sheet 4: Raw Data ---
    ws_raw = wb.create_sheet("Raw Data")
    raw_headers = ["Alternative"]
    for c in CRITERIA_ORDER:
        raw_headers.append(CRITERIA_LABELS.get(c, c))

    ws_raw.append(raw_headers)
    for col_idx in range(1, len(raw_headers) + 1):
        cell = ws_raw.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for alt in alternatives:
        row = [alt.get("name", "")]
        for c in CRITERIA_ORDER:
            row.append(alt.get(c, 0))
        ws_raw.append(row)

    for col_idx in range(1, len(raw_headers) + 1):
        ws_raw.column_dimensions[get_column_letter(col_idx)].width = 16

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ---------------------------------------------------------------------------
# High-level async generators (DB-backed)
# ---------------------------------------------------------------------------


async def generate_mcda_report(
    scenario_id: uuid.UUID,
    db: AsyncSession,
    report_format: str = "pdf",
) -> bytes | None:
    """Generate an MCDA report from a persisted scenario.

    Loads the scenario from DB, recomputes scores if needed, and
    delegates to ``generate_mcda_pdf`` or ``generate_mcda_excel``.

    Args:
        scenario_id: UUID of the MCDAScenario row.
        db: Async database session.
        report_format: ``"pdf"`` or ``"xlsx"``.

    Returns:
        Report bytes, or ``None`` if scenario not found.
    """
    scenario = await _load_scenario(scenario_id, db)
    if scenario is None:
        return None

    alts = scenario.alternatives or []
    weights = scenario.weights or dict(CDC_DEFAULT_WEIGHTS)

    # Use stored results if available, otherwise recompute
    if scenario.results and "ranked_alternatives" in scenario.results:
        results = scenario.results
    else:
        results = compute_mcda_scores(alts, weights)

    if report_format == "xlsx":
        return generate_mcda_excel(scenario.name, alts, weights, results)
    return generate_mcda_pdf(scenario.name, alts, weights, results)
