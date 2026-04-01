from __future__ import annotations

import logging
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.employee import Employee
from app.models.leave import EmployeeLeave
from app.models.site import Site

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Valid leave types (must stay in sync with app.schemas.leave)
# ---------------------------------------------------------------------------

VALID_LEAVE_TYPES = ("vacation", "sick", "unpaid", "formation", "mission", "other")


# ---------------------------------------------------------------------------
# Result dataclasses
# ---------------------------------------------------------------------------


@dataclass
class ImportError:
    """A single validation or import error tied to a specific cell."""

    sheet: str
    row: int
    column: str
    message: str


@dataclass
class SheetResult:
    """Import outcome for one sheet."""

    sheet: str
    rows_parsed: int = 0
    rows_imported: int = 0
    rows_skipped: int = 0
    errors: list[ImportError] = field(default_factory=list)


@dataclass
class ImportResult:
    """Aggregate import outcome across all processed sheets."""

    sheets: list[SheetResult] = field(default_factory=list)
    total_errors: int = 0
    is_preview: bool = False


# ---------------------------------------------------------------------------
# Sheet name constants
# ---------------------------------------------------------------------------

SHEET_SITES = "SITES"
SHEET_EFFECTIF = "EFFECTIF"
SHEET_USAGES = "USAGES & MODES"
SHEET_CONTRAINTES = "CONTRAINTES"
SHEET_PARC = "PARC EXISTANT"
SHEET_ABSENCES = "ABSENCES"

ALL_SHEETS = [
    SHEET_SITES,
    SHEET_EFFECTIF,
    SHEET_USAGES,
    SHEET_CONTRAINTES,
    SHEET_PARC,
    SHEET_ABSENCES,
]


# ---------------------------------------------------------------------------
# Main service
# ---------------------------------------------------------------------------


class ExcelImportService:
    """Parses and imports the multi-sheet Transpop Excel template."""

    def __init__(self, db: AsyncSession, tenant_id: uuid.UUID) -> None:
        self.db = db
        self.tenant_id = tenant_id
        # Caches populated lazily during parsing
        self._sites_by_code: dict[str, Site] | None = None
        self._employees_by_matricule: dict[str, Employee] | None = None

    # ------------------------------------------------------------------
    # Public entry point
    # ------------------------------------------------------------------

    async def parse_and_import(
        self,
        file_bytes: bytes,
        preview: bool = False,
        sheet_name: str | None = None,
    ) -> ImportResult:
        """Parse the Excel file.

        If *preview* is ``True``, validate only (no DB writes).
        If *sheet_name* is given, process only that sheet.
        """
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        results: list[SheetResult] = []

        sheets_to_process = [sheet_name] if sheet_name else ALL_SHEETS

        for name in sheets_to_process:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            if name == SHEET_SITES:
                results.append(await self._parse_sites(ws, preview))
                if not preview:
                    await self.db.flush()
                    self._sites_by_code = None  # invalidate cache
            elif name == SHEET_EFFECTIF:
                results.append(await self._parse_effectif(ws, preview))
                if not preview:
                    await self.db.flush()
                    self._employees_by_matricule = None  # invalidate cache
            elif name == SHEET_USAGES:
                results.append(await self._parse_usages(ws, preview))
            elif name == SHEET_CONTRAINTES:
                results.append(self._parse_contraintes(ws, preview))
            elif name == SHEET_PARC:
                results.append(self._parse_parc(ws, preview))
            elif name == SHEET_ABSENCES:
                results.append(await self._parse_absences(ws, preview))

        wb.close()

        if not preview:
            await self.db.flush()

        total_errors = sum(len(r.errors) for r in results)
        return ImportResult(
            sheets=results,
            total_errors=total_errors,
            is_preview=preview,
        )

    # ------------------------------------------------------------------
    # Site cache helpers
    # ------------------------------------------------------------------

    async def _load_sites_cache(self) -> dict[str, Site]:
        """Load or return cached sites keyed by code for the tenant."""
        if self._sites_by_code is None:
            stmt = select(Site).where(Site.tenant_id == self.tenant_id)
            result = await self.db.execute(stmt)
            self._sites_by_code = {s.code: s for s in result.scalars().all()}
        return self._sites_by_code

    async def _load_employees_cache(self) -> dict[str, Employee]:
        """Load or return cached employees keyed by matricule for the tenant."""
        if self._employees_by_matricule is None:
            stmt = select(Employee).where(Employee.tenant_id == self.tenant_id)
            result = await self.db.execute(stmt)
            self._employees_by_matricule = {
                e.matricule: e for e in result.scalars().all()
            }
        return self._employees_by_matricule

    # ------------------------------------------------------------------
    # Helper utilities
    # ------------------------------------------------------------------

    @staticmethod
    def _get_header_map(ws: Worksheet) -> dict[str, int]:
        """Read row 1 of *ws* and return a mapping of column name -> column index (1-based)."""
        header_map: dict[str, int] = {}
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
            for cell in row:
                if cell.value is not None:
                    header_map[str(cell.value).strip()] = cell.column
        return header_map

    @staticmethod
    def _cell_value(row: tuple[Any, ...], col_idx: int) -> Any:
        """Safely get cell value from a row tuple at 1-based *col_idx*.

        *row* is a tuple of Cell objects produced by ``ws.iter_rows()``.
        """
        if col_idx < 1 or col_idx > len(row):
            return None
        cell = row[col_idx - 1]
        return cell.value if hasattr(cell, "value") else cell

    @staticmethod
    def _parse_gps(value: Any) -> tuple[float, float] | None:
        """Parse a GPS string like ``"33.57, -7.59"`` into ``(lat, lng)``.

        Also accepts already-numeric types (float/int).
        Returns ``None`` if parsing fails.
        """
        if value is None:
            return None

        if isinstance(value, (int, float)):
            # Single numeric value cannot represent a coordinate pair
            return None

        text = str(value).strip()
        if not text:
            return None

        # Try comma-separated "lat, lng"
        parts = [p.strip() for p in text.split(",")]
        if len(parts) == 2:
            try:
                lat = float(parts[0])
                lng = float(parts[1])
                return (lat, lng)
            except ValueError:
                return None

        return None

    @staticmethod
    def _parse_float(value: Any) -> float | None:
        """Attempt to coerce a cell value to float."""
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).strip().replace(",", "."))
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _parse_int(value: Any) -> int | None:
        """Attempt to coerce a cell value to int."""
        if value is None:
            return None
        if isinstance(value, int):
            return value
        try:
            return int(float(str(value).strip()))
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _parse_bool(value: Any) -> bool:
        """Parse a cell value to bool (French-friendly)."""
        if value is None:
            return False
        text = str(value).strip().lower()
        return text in ("true", "1", "oui", "yes", "vrai", "x")

    @staticmethod
    def _parse_date(value: Any) -> date | None:
        """Parse a cell value to a Python date."""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value).strip()
        # Try common formats
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def _str_val(value: Any) -> str | None:
        """Convert a cell value to a stripped string or None."""
        if value is None:
            return None
        text = str(value).strip()
        return text if text else None

    @staticmethod
    def _make_geom(lng: float, lat: float) -> Any:
        """Return a PostGIS POINT expression."""
        return func.ST_SetSRID(func.ST_MakePoint(lng, lat), 4326)

    # ------------------------------------------------------------------
    # SITES sheet parser
    # ------------------------------------------------------------------

    async def _parse_sites(self, ws: Worksheet, preview: bool) -> SheetResult:
        """Parse and optionally import the SITES sheet."""
        result = SheetResult(sheet=SHEET_SITES)
        header = self._get_header_map(ws)

        # Column name mappings (French header -> internal)
        col_code = header.get("Code Site")
        col_name = header.get("Nom Site")
        col_address = header.get("Adresse")
        col_city = header.get("Ville")
        col_lat = header.get("Latitude")
        col_lng = header.get("Longitude")
        col_gps = header.get("GPS")
        col_num_shifts = header.get("Nombre Equipes")
        col_contact = header.get("Contact")
        col_phone = header.get("Telephone")
        col_zfe = header.get("ZFE")
        col_observations = header.get("Observations")

        # Load existing sites for upsert
        sites_cache = await self._load_sites_cache()

        for row_tuple in ws.iter_rows(min_row=2, values_only=False):
            row_num = row_tuple[0].row if hasattr(row_tuple[0], "row") else result.rows_parsed + 2
            result.rows_parsed += 1

            # Read values
            code = self._str_val(self._cell_value(row_tuple, col_code)) if col_code else None
            name = self._str_val(self._cell_value(row_tuple, col_name)) if col_name else None
            address = self._str_val(self._cell_value(row_tuple, col_address)) if col_address else None
            city = self._str_val(self._cell_value(row_tuple, col_city)) if col_city else None

            # Skip entirely blank rows
            if not code and not name and not address and not city:
                result.rows_skipped += 1
                continue

            # Validate required fields
            row_errors: list[ImportError] = []
            if not code:
                row_errors.append(ImportError(SHEET_SITES, row_num, "Code Site", "Required field"))
            if not name:
                row_errors.append(ImportError(SHEET_SITES, row_num, "Nom Site", "Required field"))
            if not address:
                row_errors.append(ImportError(SHEET_SITES, row_num, "Adresse", "Required field"))
            if not city:
                row_errors.append(ImportError(SHEET_SITES, row_num, "Ville", "Required field"))

            if row_errors:
                result.errors.extend(row_errors)
                result.rows_skipped += 1
                continue

            # Parse lat/lng: try separate columns first, then combined GPS column
            lat: float | None = None
            lng: float | None = None

            if col_lat and col_lng:
                lat = self._parse_float(self._cell_value(row_tuple, col_lat))
                lng = self._parse_float(self._cell_value(row_tuple, col_lng))

            if (lat is None or lng is None) and col_gps:
                gps_pair = self._parse_gps(self._cell_value(row_tuple, col_gps))
                if gps_pair:
                    lat, lng = gps_pair

            # Validate lat/lng ranges if provided
            if lat is not None and (lat < -90 or lat > 90):
                result.errors.append(
                    ImportError(SHEET_SITES, row_num, "Latitude", f"Invalid latitude: {lat}")
                )
                result.rows_skipped += 1
                continue
            if lng is not None and (lng < -180 or lng > 180):
                result.errors.append(
                    ImportError(SHEET_SITES, row_num, "Longitude", f"Invalid longitude: {lng}")
                )
                result.rows_skipped += 1
                continue

            # Sites require lat/lng (model has nullable=False)
            if lat is None or lng is None:
                result.errors.append(
                    ImportError(SHEET_SITES, row_num, "Latitude/Longitude", "GPS coordinates are required for sites")
                )
                result.rows_skipped += 1
                continue

            # Optional fields
            num_shifts = self._parse_int(self._cell_value(row_tuple, col_num_shifts)) if col_num_shifts else 1
            contact = self._str_val(self._cell_value(row_tuple, col_contact)) if col_contact else None
            phone = self._str_val(self._cell_value(row_tuple, col_phone)) if col_phone else None
            zfe = self._parse_bool(self._cell_value(row_tuple, col_zfe)) if col_zfe else False
            observations = self._str_val(self._cell_value(row_tuple, col_observations)) if col_observations else None

            if not preview:
                existing = sites_cache.get(code)  # type: ignore[arg-type]
                if existing is not None:
                    # Update existing site
                    existing.name = name  # type: ignore[assignment]
                    existing.address = address  # type: ignore[assignment]
                    existing.city = city  # type: ignore[assignment]
                    existing.lat = lat
                    existing.lng = lng
                    existing.geom = self._make_geom(lng, lat)
                    if num_shifts is not None:
                        existing.num_shifts = num_shifts
                    if contact is not None:
                        existing.contact_name = contact
                    if phone is not None:
                        existing.contact_phone = phone
                    existing.zfe_zone = zfe
                    if observations is not None:
                        existing.observations = observations
                else:
                    # Create new site
                    site = Site(
                        tenant_id=self.tenant_id,
                        code=code,  # type: ignore[arg-type]
                        name=name,  # type: ignore[arg-type]
                        address=address,  # type: ignore[arg-type]
                        city=city,  # type: ignore[arg-type]
                        lat=lat,
                        lng=lng,
                        geom=self._make_geom(lng, lat),
                        num_shifts=num_shifts or 1,
                        contact_name=contact,
                        contact_phone=phone,
                        zfe_zone=zfe,
                        observations=observations,
                    )
                    self.db.add(site)
                    sites_cache[code] = site  # type: ignore[index]

            result.rows_imported += 1

        logger.info(
            "SITES sheet: %d parsed, %d imported, %d skipped, %d errors",
            result.rows_parsed,
            result.rows_imported,
            result.rows_skipped,
            len(result.errors),
        )
        return result

    # ------------------------------------------------------------------
    # EFFECTIF sheet parser
    # ------------------------------------------------------------------

    async def _parse_effectif(self, ws: Worksheet, preview: bool) -> SheetResult:
        """Parse and optionally import the EFFECTIF (employees) sheet."""
        result = SheetResult(sheet=SHEET_EFFECTIF)
        header = self._get_header_map(ws)

        col_matricule = header.get("Matricule")
        col_prenom = header.get("Prenom")
        col_nom = header.get("Nom")
        col_site_code = header.get("Code Site")
        col_equipe = header.get("Equipe")
        col_address = header.get("Adresse")
        col_city = header.get("Ville")
        col_lat = header.get("Latitude")
        col_lng = header.get("Longitude")
        col_gps = header.get("GPS")
        col_pmr = header.get("PMR")
        col_department = header.get("Departement")
        col_phone = header.get("Telephone")
        col_mode = header.get("Mode Transport")
        col_opt_in = header.get("Opt-in")

        sites_cache = await self._load_sites_cache()
        employees_cache = await self._load_employees_cache()

        for row_tuple in ws.iter_rows(min_row=2, values_only=False):
            row_num = row_tuple[0].row if hasattr(row_tuple[0], "row") else result.rows_parsed + 2
            result.rows_parsed += 1

            matricule = self._str_val(self._cell_value(row_tuple, col_matricule)) if col_matricule else None
            prenom = self._str_val(self._cell_value(row_tuple, col_prenom)) if col_prenom else None
            nom = self._str_val(self._cell_value(row_tuple, col_nom)) if col_nom else None
            site_code = self._str_val(self._cell_value(row_tuple, col_site_code)) if col_site_code else None

            # Skip blank rows
            if not matricule and not prenom and not nom:
                result.rows_skipped += 1
                continue

            # Validate required fields
            row_errors: list[ImportError] = []
            if not matricule:
                row_errors.append(ImportError(SHEET_EFFECTIF, row_num, "Matricule", "Required field"))
            if not prenom:
                row_errors.append(ImportError(SHEET_EFFECTIF, row_num, "Prenom", "Required field"))
            if not nom:
                row_errors.append(ImportError(SHEET_EFFECTIF, row_num, "Nom", "Required field"))
            if not site_code:
                row_errors.append(ImportError(SHEET_EFFECTIF, row_num, "Code Site", "Required field"))

            if row_errors:
                result.errors.extend(row_errors)
                result.rows_skipped += 1
                continue

            # Validate site code reference
            site = sites_cache.get(site_code)  # type: ignore[arg-type]
            if site is None:
                result.errors.append(
                    ImportError(
                        SHEET_EFFECTIF,
                        row_num,
                        "Code Site",
                        f"Site with code '{site_code}' not found",
                    )
                )
                result.rows_skipped += 1
                continue

            # Parse optional fields
            equipe = self._str_val(self._cell_value(row_tuple, col_equipe)) if col_equipe else None
            address = self._str_val(self._cell_value(row_tuple, col_address)) if col_address else None
            city = self._str_val(self._cell_value(row_tuple, col_city)) if col_city else None
            department = self._str_val(self._cell_value(row_tuple, col_department)) if col_department else None
            phone = self._str_val(self._cell_value(row_tuple, col_phone)) if col_phone else None
            mode = self._str_val(self._cell_value(row_tuple, col_mode)) if col_mode else None
            opt_in_raw = self._str_val(self._cell_value(row_tuple, col_opt_in)) if col_opt_in else None
            is_pmr = self._parse_bool(self._cell_value(row_tuple, col_pmr)) if col_pmr else False

            # Parse opt-in value
            opt_in = "Non"
            if opt_in_raw:
                if opt_in_raw in ("Oui", "Non", "Sous conditions"):
                    opt_in = opt_in_raw
                elif opt_in_raw.lower() in ("oui", "yes", "true", "1"):
                    opt_in = "Oui"

            # Parse coordinates
            lat: float | None = None
            lng: float | None = None

            if col_lat and col_lng:
                lat = self._parse_float(self._cell_value(row_tuple, col_lat))
                lng = self._parse_float(self._cell_value(row_tuple, col_lng))

            if (lat is None or lng is None) and col_gps:
                gps_pair = self._parse_gps(self._cell_value(row_tuple, col_gps))
                if gps_pair:
                    lat, lng = gps_pair

            # Validate lat/lng ranges
            if lat is not None and (lat < -90 or lat > 90):
                result.errors.append(
                    ImportError(SHEET_EFFECTIF, row_num, "Latitude", f"Invalid latitude: {lat}")
                )
                result.rows_skipped += 1
                continue
            if lng is not None and (lng < -180 or lng > 180):
                result.errors.append(
                    ImportError(SHEET_EFFECTIF, row_num, "Longitude", f"Invalid longitude: {lng}")
                )
                result.rows_skipped += 1
                continue

            geom = self._make_geom(lng, lat) if (lat is not None and lng is not None) else None

            if not preview:
                existing = employees_cache.get(matricule)  # type: ignore[arg-type]
                if existing is not None:
                    # Update existing employee
                    existing.first_name = prenom  # type: ignore[assignment]
                    existing.last_name = nom  # type: ignore[assignment]
                    existing.site_id = site.id
                    existing.shift_time = equipe
                    existing.address = address
                    existing.city = city
                    existing.department = department
                    existing.phone = phone
                    existing.current_transport_mode = mode
                    existing.opt_in_company_transport = opt_in
                    existing.is_pmr = is_pmr
                    if lat is not None:
                        existing.lat = lat
                    if lng is not None:
                        existing.lng = lng
                    if geom is not None:
                        existing.geom = geom
                else:
                    # Create new employee
                    employee = Employee(
                        tenant_id=self.tenant_id,
                        matricule=matricule,  # type: ignore[arg-type]
                        first_name=prenom,  # type: ignore[arg-type]
                        last_name=nom,  # type: ignore[arg-type]
                        site_id=site.id,
                        shift_time=equipe,
                        address=address,
                        city=city,
                        lat=lat,
                        lng=lng,
                        geom=geom,
                        is_pmr=is_pmr,
                        department=department,
                        phone=phone,
                        current_transport_mode=mode,
                        opt_in_company_transport=opt_in,
                    )
                    self.db.add(employee)
                    employees_cache[matricule] = employee  # type: ignore[index]

            result.rows_imported += 1

        logger.info(
            "EFFECTIF sheet: %d parsed, %d imported, %d skipped, %d errors",
            result.rows_parsed,
            result.rows_imported,
            result.rows_skipped,
            len(result.errors),
        )
        return result

    # ------------------------------------------------------------------
    # USAGES & MODES sheet parser
    # ------------------------------------------------------------------

    async def _parse_usages(self, ws: Worksheet, preview: bool) -> SheetResult:
        """Parse and validate the USAGES & MODES sheet.

        The EmployeeModal model is not yet created (session 15), so this
        method validates data but does not write to the database.
        """
        result = SheetResult(sheet=SHEET_USAGES)
        header = self._get_header_map(ws)

        col_matricule = header.get("Matricule")
        col_mode_principal = header.get("Mode Principal")
        col_mode_alt = header.get("Mode Alternatif")
        col_distance = header.get("Distance km")
        col_temps = header.get("Temps Trajet")
        col_frequence = header.get("Frequence")
        col_point_commun = header.get("Accepte Point Commun")

        employees_cache = await self._load_employees_cache()

        for row_tuple in ws.iter_rows(min_row=2, values_only=False):
            row_num = row_tuple[0].row if hasattr(row_tuple[0], "row") else result.rows_parsed + 2
            result.rows_parsed += 1

            matricule = self._str_val(self._cell_value(row_tuple, col_matricule)) if col_matricule else None

            # Skip blank rows
            if not matricule:
                result.rows_skipped += 1
                continue

            # Validate employee reference
            if matricule not in employees_cache:
                result.errors.append(
                    ImportError(
                        SHEET_USAGES,
                        row_num,
                        "Matricule",
                        f"Employee with matricule '{matricule}' not found",
                    )
                )
                result.rows_skipped += 1
                continue

            # Validate distance if provided
            if col_distance:
                distance_raw = self._cell_value(row_tuple, col_distance)
                if distance_raw is not None:
                    distance = self._parse_float(distance_raw)
                    if distance is None:
                        result.errors.append(
                            ImportError(
                                SHEET_USAGES,
                                row_num,
                                "Distance km",
                                f"Invalid numeric value: {distance_raw}",
                            )
                        )
                        result.rows_skipped += 1
                        continue

            # Data validated; DB write skipped (EmployeeModal not yet created)
            result.rows_imported += 1

        logger.info(
            "USAGES & MODES sheet: %d parsed, %d validated, %d skipped, %d errors",
            result.rows_parsed,
            result.rows_imported,
            result.rows_skipped,
            len(result.errors),
        )
        return result

    # ------------------------------------------------------------------
    # CONTRAINTES sheet parser
    # ------------------------------------------------------------------

    def _parse_contraintes(self, ws: Worksheet, preview: bool) -> SheetResult:
        """Parse and validate the CONTRAINTES sheet.

        The Constraint model is not yet created (session 29), so this
        method validates key-value format but does not write to the database.
        """
        result = SheetResult(sheet=SHEET_CONTRAINTES)
        header = self._get_header_map(ws)

        col_key = header.get("Parametre") or header.get("Cle") or header.get("Key")
        col_value = header.get("Valeur") or header.get("Value")

        for row_tuple in ws.iter_rows(min_row=2, values_only=False):
            row_num = row_tuple[0].row if hasattr(row_tuple[0], "row") else result.rows_parsed + 2
            result.rows_parsed += 1

            key = self._str_val(self._cell_value(row_tuple, col_key)) if col_key else None
            value = self._str_val(self._cell_value(row_tuple, col_value)) if col_value else None

            # Skip blank rows
            if not key and not value:
                result.rows_skipped += 1
                continue

            if not key:
                result.errors.append(
                    ImportError(SHEET_CONTRAINTES, row_num, "Parametre", "Required field")
                )
                result.rows_skipped += 1
                continue

            # Validated; DB write deferred to session 29
            result.rows_imported += 1

        logger.info(
            "CONTRAINTES sheet: %d parsed, %d validated, %d skipped, %d errors",
            result.rows_parsed,
            result.rows_imported,
            result.rows_skipped,
            len(result.errors),
        )
        return result

    # ------------------------------------------------------------------
    # PARC EXISTANT sheet parser
    # ------------------------------------------------------------------

    def _parse_parc(self, ws: Worksheet, preview: bool) -> SheetResult:
        """Parse and validate the PARC EXISTANT (vehicle fleet) sheet.

        The Vehicle model is not yet created (session 20), so this
        method validates data but does not write to the database.
        """
        result = SheetResult(sheet=SHEET_PARC)
        header = self._get_header_map(ws)

        col_type = header.get("Type Vehicule") or header.get("Type")
        col_capacity = header.get("Capacite") or header.get("Places")
        col_immat = header.get("Immatriculation")

        for row_tuple in ws.iter_rows(min_row=2, values_only=False):
            row_num = row_tuple[0].row if hasattr(row_tuple[0], "row") else result.rows_parsed + 2
            result.rows_parsed += 1

            vtype = self._str_val(self._cell_value(row_tuple, col_type)) if col_type else None
            capacity_raw = self._cell_value(row_tuple, col_capacity) if col_capacity else None

            # Skip blank rows
            if not vtype and capacity_raw is None:
                result.rows_skipped += 1
                continue

            # Validate capacity is numeric
            if capacity_raw is not None:
                capacity = self._parse_int(capacity_raw)
                if capacity is None:
                    result.errors.append(
                        ImportError(
                            SHEET_PARC,
                            row_num,
                            "Capacite",
                            f"Invalid numeric value: {capacity_raw}",
                        )
                    )
                    result.rows_skipped += 1
                    continue

            # Validated; DB write deferred to session 20
            result.rows_imported += 1

        logger.info(
            "PARC EXISTANT sheet: %d parsed, %d validated, %d skipped, %d errors",
            result.rows_parsed,
            result.rows_imported,
            result.rows_skipped,
            len(result.errors),
        )
        return result

    # ------------------------------------------------------------------
    # ABSENCES sheet parser
    # ------------------------------------------------------------------

    async def _parse_absences(self, ws: Worksheet, preview: bool) -> SheetResult:
        """Parse and optionally import the ABSENCES sheet."""
        result = SheetResult(sheet=SHEET_ABSENCES)
        header = self._get_header_map(ws)

        col_matricule = header.get("Matricule")
        col_type = header.get("Type Absence")
        col_start = header.get("Date Debut")
        col_end = header.get("Date Fin")

        employees_cache = await self._load_employees_cache()

        for row_tuple in ws.iter_rows(min_row=2, values_only=False):
            row_num = row_tuple[0].row if hasattr(row_tuple[0], "row") else result.rows_parsed + 2
            result.rows_parsed += 1

            matricule = self._str_val(self._cell_value(row_tuple, col_matricule)) if col_matricule else None
            leave_type = self._str_val(self._cell_value(row_tuple, col_type)) if col_type else None

            # Skip blank rows
            if not matricule and not leave_type:
                result.rows_skipped += 1
                continue

            # Validate required fields
            row_errors: list[ImportError] = []
            if not matricule:
                row_errors.append(ImportError(SHEET_ABSENCES, row_num, "Matricule", "Required field"))
            if not leave_type:
                row_errors.append(ImportError(SHEET_ABSENCES, row_num, "Type Absence", "Required field"))

            if row_errors:
                result.errors.extend(row_errors)
                result.rows_skipped += 1
                continue

            # Validate employee reference
            employee = employees_cache.get(matricule)  # type: ignore[arg-type]
            if employee is None:
                result.errors.append(
                    ImportError(
                        SHEET_ABSENCES,
                        row_num,
                        "Matricule",
                        f"Employee with matricule '{matricule}' not found",
                    )
                )
                result.rows_skipped += 1
                continue

            # Validate leave type
            if leave_type not in VALID_LEAVE_TYPES:
                result.errors.append(
                    ImportError(
                        SHEET_ABSENCES,
                        row_num,
                        "Type Absence",
                        f"Invalid leave type '{leave_type}'. Must be one of: {', '.join(VALID_LEAVE_TYPES)}",
                    )
                )
                result.rows_skipped += 1
                continue

            # Parse dates
            start_date_raw = self._cell_value(row_tuple, col_start) if col_start else None
            end_date_raw = self._cell_value(row_tuple, col_end) if col_end else None

            start_date = self._parse_date(start_date_raw)
            end_date = self._parse_date(end_date_raw)

            if start_date is None:
                result.errors.append(
                    ImportError(SHEET_ABSENCES, row_num, "Date Debut", f"Invalid or missing date: {start_date_raw}")
                )
                result.rows_skipped += 1
                continue
            if end_date is None:
                result.errors.append(
                    ImportError(SHEET_ABSENCES, row_num, "Date Fin", f"Invalid or missing date: {end_date_raw}")
                )
                result.rows_skipped += 1
                continue

            if end_date < start_date:
                result.errors.append(
                    ImportError(
                        SHEET_ABSENCES,
                        row_num,
                        "Date Fin",
                        "End date must be >= start date",
                    )
                )
                result.rows_skipped += 1
                continue

            if not preview:
                leave = EmployeeLeave(
                    employee_id=employee.id,
                    leave_type=leave_type,  # type: ignore[arg-type]
                    start_date=start_date,
                    end_date=end_date,
                )
                self.db.add(leave)

            result.rows_imported += 1

        logger.info(
            "ABSENCES sheet: %d parsed, %d imported, %d skipped, %d errors",
            result.rows_parsed,
            result.rows_imported,
            result.rows_skipped,
            len(result.errors),
        )
        return result
